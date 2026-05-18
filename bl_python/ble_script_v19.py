#!/usr/bin/env python3
"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║         ⚡ VPUSH BMS DASHBOARD v5.2 — INTEGRATED OTA (FULLY FIXED)           ║
║                                                                               ║
║  OTA FIXES vs v5.1:                                                          ║
║    • Complete OTA engine rewrite to match ota.py v3.0 EXACTLY                ║
║    • Proper asyncio event handling for ACKs                                  ║
║    • Fixed notification handler to set OTA events on correct loop            ║
║    • Added missing OTA state machine reset                                   ║
║    • CRC-32: reflected poly 0xEDB88320 matches bsp.c                         ║
║                                                                               ║
║  Frame: [0x56][0x50][DIR][CMD][LEN_H][LEN_L][PAYLOAD…][CS][0xAA][0xAA]       ║
║  CS = ~(CMD + sum(PAYLOAD)) & 0xFF                                            ║
║                                                                               ║
║  OTA SEQUENCE (fully automatic, matches ota.py v3.0):                        ║
║    Step 1: TRIGGER   (0x36)                                                   ║
║    Step 2: FLASH_CMD (0x33) → erases application flash                       ║
║    Step 3: FILE_SIZE (0x01) → sends padded image size                        ║
║    Step 4: CHUNK     (0x02) × N → 232 B chunks                               ║
║    Step 5: CRC       (0x03) → CRC-32 verify, device reboots                  ║
║                                                                               ║
║  Request UUID : b7453898-cbc6-4209-8a19-e804e76042d7                         ║
║  Notify  UUID : d0f57f59-6f8f-4773-8068-7a8a5b62a86f                        ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""

import asyncio
import threading
import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox
from bleak import BleakScanner, BleakClient
from datetime import datetime
import traceback
import os
import struct
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    import pandas as pd
    _PANDAS_OK = True
except ImportError:
    _PANDAS_OK = False

# =============================================================================
# PROTOCOL CONSTANTS
# =============================================================================

NOTIFY_UUID  = "d0f57f59-6f8f-4773-8068-7a8a5b62a86f"
REQUEST_UUID = "b7453898-cbc6-4209-8a19-e804e76042d7"
DEFAULT_DEVICE_NAME = "Thunderboard"

SOF_BYTE_0      = 0x56
SOF_BYTE_1      = 0x50
EOF_BYTE_0      = 0xAA
EOF_BYTE_1      = 0xAA
HOST_TO_BLE_DIR = 0x00
BLE_TO_HOST_DIR = 0x01

CMD_BAT_START  = 0x01
CMD_BAT_STOP   = 0x02
CMD_SD_DATA    = 0x03
CMD_BMS_CONFIG = 0x04
CMD_OTA_UPDATE = 0x05

SD_CARD_INFO  = 0xAA
SD_READ_CARD  = 0xAB
SD_ERASE_CARD = 0xAC

# OTA sub-commands (identical to ota.py v3.0)
OTA_SUBCMD_TRIGGER   = 0x36
OTA_SUBCMD_FLASH_CMD = 0x33
OTA_SUBCMD_FILE_SIZE = 0x01
OTA_SUBCMD_CHUNK     = 0x02
OTA_SUBCMD_CRC       = 0x03
OTA_ACK_OK           = 0x01
OTA_ACK_FAIL         = 0x00

CHUNK_SIZE      = 232
OTA_ACK_TIMEOUT = 8.0
OTA_MAX_RETRIES = 3

# BMS config
BMS_CFG_TYPE_U32     = 0xAA
BMS_CFG_TYPE_I32     = 0xBB
BMS_CFG_TYPE_I16     = 0xCC
FLASH_BASE_ADDRESS   = 0x10000000
BMS_CFG_ACK_TIMEOUT  = 5.0
BMS_CFG_MAX_RETRIES  = 3

COMMAND_INTERVAL        = 1.0
SUBFRAMES_PER_SECTOR    = 4
SD_RESPONSE_TIMEOUT_SEC = 60.0


# =============================================================================
# CRC-32 — reflected poly 0xEDB88320 (matches BSP_Crc32_u32 in bsp.c)
# =============================================================================

def crc32_firmware(data: bytes) -> int:
    crc = 0xFFFFFFFF
    for byte in data:
        crc ^= byte
        for _ in range(8):
            if crc & 1:
                crc = (crc >> 1) ^ 0xEDB88320
            else:
                crc >>= 1
    return (~crc) & 0xFFFFFFFF


# =============================================================================
# FRAME BUILDER - EXACT MATCH to ota.py v3.0
# =============================================================================

def _checksum(cmd: int, payload: bytes) -> int:
    """One's-complement checksum over cmd + all payload bytes."""
    s = cmd + sum(payload)
    return 0xFF & (~s & 0xFF)


def build_frame(cmd_id: int, payload: bytes = b"") -> bytearray:
    """
    Build a host-to-device frame WITH checksum — identical to ota.py v3.0.
    [SOF0][SOF1][DIR_H2D][CMD][LEN_H][LEN_L][PAYLOAD…][CS][EOF0][EOF1]
    """
    plen = len(payload)
    cs   = _checksum(cmd_id, payload)
    frame = bytearray([
        SOF_BYTE_0, SOF_BYTE_1,
        HOST_TO_BLE_DIR,
        cmd_id,
        (plen >> 8) & 0xFF,
        plen & 0xFF,
    ])
    frame.extend(payload)
    frame.extend([cs, EOF_BYTE_0, EOF_BYTE_1])
    return frame


def build_bms_config_payload(address: int, data_type_str: str, value) -> bytes:
    offset       = address - FLASH_BASE_ADDRESS
    offset_bytes = struct.pack('<H', offset & 0xFFFF)
    length_byte  = bytes([4])
    dt = data_type_str.upper()
    if dt == 'I32':
        type_byte = bytes([BMS_CFG_TYPE_I32])
        val_bytes = struct.pack('<f', float(value))
    elif dt in ('I16', 'I8'):
        type_byte = bytes([BMS_CFG_TYPE_I16])
        val_bytes = struct.pack('<i', int(value))[:4]
    else:
        type_byte = bytes([BMS_CFG_TYPE_U32])
        val_bytes = struct.pack('<I', int(value) & 0xFFFFFFFF)
    return offset_bytes + length_byte + type_byte + val_bytes


# =============================================================================
# BMS SIGNAL DECODER
# =============================================================================

class BMSSignalDecoder:

    @staticmethod
    def decode(p):
        if len(p) < 100:
            return None
        s = {}
        o = 0
        try:
            s['TimeStamp'] = p[o]|(p[o+1]<<8)|(p[o+2]<<16)|(p[o+3]<<24); o+=4
            r = p[o]|(p[o+1]<<8)
            if r & 0x8000: r = -((~r+1) & 0xFFFF)
            s['PackCurrent_A'] = round(r / 100.0, 2); o+=2
            s['PackVoltage_V'] = round((p[o]|(p[o+1]<<8)) / 10.0, 2); o+=2
            for i in range(16):
                mv = p[o]|(p[o+1]<<8)
                s[f'Cell{i+1}_V'] = round(mv / 1000.0, 3); o+=2
            s['MaxCellVoltage_V'] = round((p[o]|(p[o+1]<<8)) / 1000.0, 3); o+=2
            s['MinCellVoltage_V'] = round((p[o]|(p[o+1]<<8)) / 1000.0, 3); o+=2
            s['BattVoltageAfterFET_V'] = round((p[o]|(p[o+1]<<8)) / 1000.0, 3); o+=2
            s['LoadVoltage_V'] = round((p[o]|(p[o+1]<<8)) / 10.0, 2); o+=2
            s['CellVoltageDiff_mV'] = p[o]|(p[o+1]<<8); o+=2
            s['ChargeCapacity_Ah'] = round((p[o]|(p[o+1]<<8)) / 100.0, 2); o+=2
            s['DischargeCapacity_Ah'] = round((p[o]|(p[o+1]<<8)) / 100.0, 2); o+=2
            for i in range(5):
                raw = p[o]|(p[o+1]<<8)
                if raw & 0x8000: raw = -((~raw + 1) & 0xFFFF)
                s[f'BattTemp_{i+1}_C'] = round(raw / 100.0, 1); o+=2
            for i in range(5):
                raw = p[o]|(p[o+1]<<8)
                if raw & 0x8000: raw = -((~raw + 1) & 0xFFFF)
                s[f'PCBTemp_{i+1}_C'] = round(raw / 100.0, 1); o+=2
            raw = p[o]|(p[o+1]<<8)
            if raw & 0x8000: raw = -((~raw + 1) & 0xFFFF)
            s['MinBattTemp_C'] = round(raw / 100.0, 1); o+=2
            raw = p[o]|(p[o+1]<<8)
            if raw & 0x8000: raw = -((~raw + 1) & 0xFFFF)
            s['MaxBattTemp_C'] = round(raw / 100.0, 1); o+=2
            s['CellBalancingStatus'] = p[o]|(p[o+1]<<8); o+=2
            raw = p[o]|(p[o+1]<<8)
            if raw & 0x8000: raw = -((~raw + 1) & 0xFFFF)
            s['CellTempDiff_C'] = round(raw / 100.0, 1); o+=2
            s['StackVoltage_V'] = round((p[o]|(p[o+1]<<8)) / 10.0, 3); o+=2
            o += 2
            s['BMSOperationState'] = p[o]; o+=1
            fet = p[o]
            s['FETStatus'] = f"0x{fet:02X}"; o+=1
            s['SOC_percent'] = p[o]; o+=1
            s['SOH_percent'] = p[o]; o+=1
            s['SOP_percent'] = p[o]; o+=1
            s['SOCDiff_percent'] = p[o]; o+=1
            binary = p[o]
            s['BatteryChgFlag'] = bool(binary & 0x01)
            s['KeyOn']          = bool(binary & 0x02)
            s['BCUCommStatus']  = bool(binary & 0x04); o+=1
            o+=1
            f1 = p[o]
            s['MinCellUVFault']       = (f1 >> 0) & 0x03
            s['MaxCellOVFault']       = (f1 >> 2) & 0x03
            s['OverChgCurrentFault']  = (f1 >> 4) & 0x03
            s['OverDchgCurrentFault'] = (f1 >> 6) & 0x03; o+=1
            f2 = p[o]
            s['BattUTFault'] = (f2 >> 0) & 0x03
            s['BattOTFault'] = (f2 >> 2) & 0x03
            s['BattOVFault'] = (f2 >> 4) & 0x03
            s['BattUVFault'] = (f2 >> 6) & 0x03; o+=1
            tr = p[o]
            s['Thermal_MinCellUV']            = bool(tr & (1<<0))
            s['Thermal_MaxCellTemp']          = bool(tr & (1<<1))
            s['Thermal_MaxCellTempRise']      = bool(tr & (1<<2))
            s['Thermal_CellTempDiff']         = bool(tr & (1<<3))
            s['Thermal_MaxCellOV']            = bool(tr & (1<<4))
            s['Thermal_CellVolSamplingFault'] = bool(tr & (1<<5))
            s['Thermal_TempSamplingFault']    = bool(tr & (1<<6)); o+=1
            f3 = p[o]
            s['Alarm_OverPCBTemp']      = bool(f3 & (1<<0))
            s['Alarm_PreChargeFailure'] = bool(f3 & (1<<1))
            s['Alarm_HardwareSC']       = bool(f3 & (1<<2))
            s['Alarm_CellDVAlarm']      = bool(f3 & (1<<3))
            s['Alarm_CellTDAlarm']      = bool(f3 & (1<<4))
            s['Alarm_LowSOCAlarm']      = bool(f3 & (1<<5))
            s['Alarm_LowSOHAlarm']      = bool(f3 & (1<<6))
            s['Alarm_SOCJumpAlarm']     = bool(f3 & (1<<7)); o+=1
            s['ThermalRunawayFault'] = bool(p[o]); o+=1
            return s
        except Exception as e:
            print(f"[DECODE] Error at offset {o}: {e}")
            return None

    @staticmethod
    def get_excel_headers():
        h = ['Timestamp','Record#','TimeStamp(ms)','PackCurrent(A)','PackVoltage(V)',
             'MaxCellVoltage(V)','MinCellVoltage(V)','BattVoltageAfterFET(V)',
             'LoadVoltage(V)','CellVoltageDiff(mV)','ChargeCapacity(Ah)',
             'DischargeCapacity(Ah)','MinBattTemp(C)','MaxBattTemp(C)',
             'CellBalancingStatus','CellTempDiff(C)','StackVoltage(V)',
             'BMSOperationState','FETStatus','SOC(%)','SOH(%)','SOP(%)','SOCDiff(%)',
             'BatteryChgFlag','KeyOn','BCUCommStatus',
             'MinCellUVFault','MaxCellOVFault','OverChgCurrentFault','OverDchgCurrentFault',
             'BattUTFault','BattOTFault','BattOVFault','BattUVFault',
             'Thermal_MinCellUV','Thermal_MaxCellTemp','Thermal_MaxCellTempRise',
             'Thermal_CellTempDiff','Thermal_MaxCellOV','Thermal_CellVolSamplingFault',
             'Thermal_TempSamplingFault','Alarm_OverPCBTemp','Alarm_PreChargeFailure',
             'Alarm_HardwareSC','Alarm_CellDVAlarm','Alarm_CellTDAlarm',
             'Alarm_LowSOCAlarm','Alarm_LowSOHAlarm','Alarm_SOCJumpAlarm',
             'ThermalRunawayFault']
        for i in range(16): h.append(f'Cell{i+1}(V)')
        for i in range(5):  h.append(f'BattTemp_S{i+1}(C)')
        for i in range(5):  h.append(f'PCBTemp_S{i+1}(C)')
        return h

    @staticmethod
    def signals_to_row(signals, record_num, timestamp):
        b = lambda k: 1 if signals.get(k, False) else 0
        row = [timestamp, record_num,
               signals.get('TimeStamp', 0), signals.get('PackCurrent_A', 0.0),
               signals.get('PackVoltage_V', 0.0), signals.get('MaxCellVoltage_V', 0.0),
               signals.get('MinCellVoltage_V', 0.0), signals.get('BattVoltageAfterFET_V', 0.0),
               signals.get('LoadVoltage_V', 0.0), signals.get('CellVoltageDiff_mV', 0),
               signals.get('ChargeCapacity_Ah', 0.0), signals.get('DischargeCapacity_Ah', 0.0),
               signals.get('MinBattTemp_C', 0.0), signals.get('MaxBattTemp_C', 0.0),
               signals.get('CellBalancingStatus', 0), signals.get('CellTempDiff_C', 0.0),
               signals.get('StackVoltage_V', 0.0), signals.get('BMSOperationState', 0),
               signals.get('FETStatus', '0x00'), signals.get('SOC_percent', 0),
               signals.get('SOH_percent', 0), signals.get('SOP_percent', 0),
               signals.get('SOCDiff_percent', 0),
               b('BatteryChgFlag'), b('KeyOn'), b('BCUCommStatus'),
               signals.get('MinCellUVFault', 0), signals.get('MaxCellOVFault', 0),
               signals.get('OverChgCurrentFault', 0), signals.get('OverDchgCurrentFault', 0),
               signals.get('BattUTFault', 0), signals.get('BattOTFault', 0),
               signals.get('BattOVFault', 0), signals.get('BattUVFault', 0),
               b('Thermal_MinCellUV'), b('Thermal_MaxCellTemp'),
               b('Thermal_MaxCellTempRise'), b('Thermal_CellTempDiff'),
               b('Thermal_MaxCellOV'), b('Thermal_CellVolSamplingFault'),
               b('Thermal_TempSamplingFault'),
               b('Alarm_OverPCBTemp'), b('Alarm_PreChargeFailure'),
               b('Alarm_HardwareSC'), b('Alarm_CellDVAlarm'), b('Alarm_CellTDAlarm'),
               b('Alarm_LowSOCAlarm'), b('Alarm_LowSOHAlarm'), b('Alarm_SOCJumpAlarm'),
               b('ThermalRunawayFault')]
        for i in range(16): row.append(signals.get(f'Cell{i+1}_V', 0.0))
        for i in range(5):  row.append(signals.get(f'BattTemp_{i+1}_C', 0.0))
        for i in range(5):  row.append(signals.get(f'PCBTemp_{i+1}_C', 0.0))
        return row


# =============================================================================
# BATTERY DATA LOGGER
# =============================================================================

class BatteryDataLogger:
    def __init__(self):
        self.logging_active = False
        self.file_path      = None
        self.start_time     = None
        self.log_count      = 0
        self.workbook       = None
        self.sheet          = None

    def start_logging(self, directory=None, filename=None):
        if self.logging_active: return False, "Logging already active"
        try:
            if not filename:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"BMS_Live_{ts}.xlsx"
            if not directory:
                directory = filedialog.askdirectory(title="Select Log Directory")
                if not directory: return False, "No directory selected"
            self.file_path      = os.path.join(directory, filename)
            self.workbook       = Workbook()
            self.sheet          = self.workbook.active
            self.sheet.title    = "BMS Live Data"
            self._write_headers()
            self.logging_active = True
            self.start_time     = datetime.now()
            self.log_count      = 0
            return True, f"Logging started: {self.file_path}"
        except Exception as e:
            return False, f"Error: {e}"

    def stop_logging(self):
        if not self.logging_active: return False, "Logging not active"
        try:
            self.workbook.save(self.file_path)
            self.logging_active = False
            dur = datetime.now() - self.start_time
            return True, f"Saved {self.log_count} records ({dur})"
        except Exception as e:
            return False, f"Error saving: {e}"

    def _write_headers(self):
        headers = BMSSignalDecoder.get_excel_headers()
        for col, h in enumerate(headers, 1):
            cell           = self.sheet.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        self.sheet.freeze_panes = 'A2'

    def log_data(self, data):
        if not self.logging_active: return
        try:
            self.log_count += 1
            row = self.log_count + 1
            ts  = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            col = 1
            def w(v):
                nonlocal col
                self.sheet.cell(row=row, column=col, value=v); col+=1
            w(ts); w(data.packet_count)
            w(data.TimeStamp); w(round(data.PackCurrent, 2))
            w(round(data.PackVoltage, 2)); w(round(data.MaxCellVoltage, 3))
            w(round(data.MinCellVoltage, 3)); w(round(data.BattVoltageAfterFET, 3))
            w(round(data.LoadVoltage, 2)); w(data.CellVoltageDiff)
            w(round(data.ChargeCapacity, 2)); w(round(data.DischargeCapacity, 2))
            w(round(data.MinBattTemp, 1)); w(round(data.MaxBattTemp, 1))
            w(f"0x{data.CellBalancingStatus:04X}"); w(round(data.CellTempDiff, 1))
            w(round(data.StackVoltage, 3)); w(data.BMSOperationState)
            w(f"0x{data.FETStatus:02X}"); w(data.SOC)
            w(data.SOH); w(data.SOP); w(data.SOCDiff)
            w(f"0x{data.BinaryFlags:02X}"); w(f"0x{data.Reserved3:02X}")
            w(f"0x{data.Fault1:02X}"); w(f"0x{data.Fault2:02X}")
            w(f"0x{data.ThermalReason:02X}"); w(f"0x{data.Fault3:02X}")
            w(1 if data.ThermalRunawayFault else 0)
            w(f"0x{data.Reserved4:02X}")
            w(1 if data.BatteryChgFlag else 0)
            w(1 if data.KeyOn else 0)
            w(1 if data.BCUCommStatus else 0)
            for i in range(16): w(round(data.CellVoltages[i], 3))
            for i in range(5):  w(round(data.BattTemperatures[i], 1))
            for i in range(5):  w(round(data.PCBTemperatures[i], 1))
        except Exception as e:
            print(f"[BatteryDataLogger] error: {e}")


# =============================================================================
# SD CARD EXCEL LOGGER
# =============================================================================

class SDExcelLogger:
    def __init__(self):
        self.logging_active = False
        self.workbook       = None
        self.sheet          = None
        self.file_path      = None
        self.start_time     = None
        self.record_count   = 0
        self.headers        = BMSSignalDecoder.get_excel_headers()
        self._lock          = threading.Lock()

    def start_logging(self, directory, filename=None):
        if self.logging_active: return False, "Already active"
        try:
            if not filename:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"BMS_SD_{ts}.xlsx"
            self.file_path      = os.path.join(directory, filename)
            self.workbook       = Workbook()
            self.sheet          = self.workbook.active
            self.sheet.title    = "SD Decoded"
            self._write_headers()
            self.workbook.save(self.file_path)
            self.logging_active = True
            self.start_time     = datetime.now()
            self.record_count   = 0
            return True, f"SD logging started: {self.file_path}"
        except Exception as e:
            return False, f"Error: {e}"

    def stop_logging(self):
        if not self.logging_active: return False, "Not active"
        try:
            with self._lock:
                self.workbook.save(self.file_path)
                self.logging_active = False
            dur = datetime.now() - self.start_time
            return True, f"Saved {self.record_count} SD records ({dur})"
        except Exception as e:
            return False, f"Error saving: {e}"

    def _write_headers(self):
        for col, h in enumerate(self.headers, 1):
            cell           = self.sheet.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill(start_color="1F6B3A", end_color="1F6B3A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        self.sheet.freeze_panes = 'A2'

    def log_record(self, signals, record_num):
        if not self.logging_active: return False
        try:
            with self._lock:
                self.record_count += 1
                ts       = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                row_data = BMSSignalDecoder.signals_to_row(signals, record_num, ts)
                for col, value in enumerate(row_data, 1):
                    self.sheet.cell(row=record_num+1, column=col, value=value)
                self.workbook.save(self.file_path)
            return True
        except Exception as e:
            print(f"[SDExcelLogger] Error: {e}")
            return False


# =============================================================================
# BATTERY DATA CONTAINER
# =============================================================================

class BatteryData:
    __slots__ = [
        'TimeStamp','PackCurrent','PackVoltage','CellVoltages','CellVoltage_mV',
        'MaxCellVoltage','MinCellVoltage','BattVoltageAfterFET','LoadVoltage',
        'CellVoltageDiff','ChargeCapacity','DischargeCapacity','BattTemperatures',
        'PCBTemperatures','MinBattTemp','MaxBattTemp','CellBalancingStatus',
        'CellTempDiff','StackVoltage','Reserved1','Reserved2','BMSOperationState',
        'FETStatus','SOC','SOH','SOP','SOCDiff','BinaryFlags','Reserved3',
        'Fault1','Fault2','ThermalReason','Fault3','ThermalRunawayFault',
        'Reserved4','MinCellUVFault','MaxCellOVFault','OverChgCurrentFault',
        'OverDchgCurrentFault','BattUTFault','BattOTFault','BattOVFault',
        'BattUVFault','ThermalRunawayReasons','Alarms','FET_Status','connected',
        'notification_count','last_update','last_command','packet_count',
        'BatteryChgFlag','KeyOn','BCUCommStatus'
    ]

    def __init__(self):
        self.TimeStamp=0; self.PackCurrent=0.0; self.PackVoltage=0.0
        self.CellVoltages=[0.0]*16; self.CellVoltage_mV=[0]*16
        self.MaxCellVoltage=0.0; self.MinCellVoltage=0.0
        self.BattVoltageAfterFET=0.0; self.LoadVoltage=0.0
        self.CellVoltageDiff=0; self.ChargeCapacity=0.0; self.DischargeCapacity=0.0
        self.BattTemperatures=[0.0]*5; self.PCBTemperatures=[0.0]*5
        self.MinBattTemp=0.0; self.MaxBattTemp=0.0
        self.CellBalancingStatus=0; self.CellTempDiff=0.0; self.StackVoltage=0.0
        self.Reserved1=0xFF; self.Reserved2=0xFF
        self.BMSOperationState=0; self.FETStatus=0; self.SOC=0; self.SOH=0
        self.SOP=0; self.SOCDiff=0; self.BinaryFlags=0; self.Reserved3=0xFF
        self.Fault1=0; self.Fault2=0; self.ThermalReason=0; self.Fault3=0
        self.ThermalRunawayFault=False; self.Reserved4=0xFF
        self.MinCellUVFault=0; self.MaxCellOVFault=0; self.OverChgCurrentFault=0
        self.OverDchgCurrentFault=0; self.BattUTFault=0; self.BattOTFault=0
        self.BattOVFault=0; self.BattUVFault=0
        self.ThermalRunawayReasons={k:False for k in
            ['MinCellUV','MaxCellTemp','MaxCellTempRise','CellTempDiff',
             'MaxCellOV','CellVolSamplingFault','TempSamplingFault']}
        self.Alarms={k:False for k in
            ['OverPCBTemp','PreChargeFailure','HardwareSC','CellDVAlarm',
             'CellTDAlarm','LowSOCAlarm','LowSOHAlarm','SOCJumpAlarm']}
        self.FET_Status=0; self.BatteryChgFlag=False; self.KeyOn=False
        self.BCUCommStatus=False; self.connected=False
        self.notification_count=0; self.last_update="Never"
        self.last_command=0; self.packet_count=0

    def get_operation_state_string(self):
        return {0:"Idle",1:"Charging",2:"Discharging",3:"Balancing",
                4:"Fault",5:"Precharge",6:"Sleep"}.get(self.BMSOperationState,
                f"State {self.BMSOperationState}")

    def get_fault_level_string(self, level):
        return {0:"OK",1:"Warning",2:"Fault",3:"Critical"}.get(level,"Unknown")

    def get_fault_color(self, level):
        return ('#00ff95','#ffaa00','#ff4d4d','#ff4d4d')[min(level,3)] if level>=0 else '#8899aa'

    def parse_fault_bytes(self):
        self.MinCellUVFault=(self.Fault1>>0)&3; self.MaxCellOVFault=(self.Fault1>>2)&3
        self.OverChgCurrentFault=(self.Fault1>>4)&3; self.OverDchgCurrentFault=(self.Fault1>>6)&3
        self.BattUTFault=(self.Fault2>>0)&3; self.BattOTFault=(self.Fault2>>2)&3
        self.BattOVFault=(self.Fault2>>4)&3; self.BattUVFault=(self.Fault2>>6)&3
        self.ThermalRunawayReasons['MinCellUV']=bool(self.ThermalReason&1)
        self.ThermalRunawayReasons['MaxCellTemp']=bool(self.ThermalReason&2)
        self.ThermalRunawayReasons['MaxCellTempRise']=bool(self.ThermalReason&4)
        self.ThermalRunawayReasons['CellTempDiff']=bool(self.ThermalReason&8)
        self.ThermalRunawayReasons['MaxCellOV']=bool(self.ThermalReason&16)
        self.ThermalRunawayReasons['CellVolSamplingFault']=bool(self.ThermalReason&32)
        self.ThermalRunawayReasons['TempSamplingFault']=bool(self.ThermalReason&64)
        self.Alarms['OverPCBTemp']=bool(self.Fault3&1); self.Alarms['PreChargeFailure']=bool(self.Fault3&2)
        self.Alarms['HardwareSC']=bool(self.Fault3&4); self.Alarms['CellDVAlarm']=bool(self.Fault3&8)
        self.Alarms['CellTDAlarm']=bool(self.Fault3&16); self.Alarms['LowSOCAlarm']=bool(self.Fault3&32)
        self.Alarms['LowSOHAlarm']=bool(self.Fault3&64); self.Alarms['SOCJumpAlarm']=bool(self.Fault3&128)
        self.BatteryChgFlag=bool(self.FETStatus&0x01)
        self.KeyOn=bool(self.FETStatus&0x02)
        self.BCUCommStatus=bool(self.FETStatus&0x04)

    def update_from_payload(self, data):
        o=0; total=len(data)
        try:
            self.TimeStamp=(data[o]|(data[o+1]<<8)|(data[o+2]<<16)|(data[o+3]<<24)); o+=4
            r=data[o]|(data[o+1]<<8)
            if r&0x8000: r=-((~r+1)&0xFFFF)
            self.PackCurrent=r/100.0; o+=2
            self.PackVoltage=(data[o]|(data[o+1]<<8))/100.0; o+=2
            for i in range(16):
                mv=data[o]|(data[o+1]<<8)
                self.CellVoltage_mV[i]=mv; self.CellVoltages[i]=mv/1000.0; o+=2
            self.MaxCellVoltage=(data[o]|(data[o+1]<<8))/1000.0; o+=2
            self.MinCellVoltage=(data[o]|(data[o+1]<<8))/1000.0; o+=2
            self.BattVoltageAfterFET=(data[o]|(data[o+1]<<8))/10.0; o+=2
            self.LoadVoltage=(data[o]|(data[o+1]<<8))/100.0; o+=2
            self.CellVoltageDiff=data[o]|(data[o+1]<<8); o+=2
            self.ChargeCapacity=(data[o]|(data[o+1]<<8))/100.0; o+=2
            self.DischargeCapacity=(data[o]|(data[o+1]<<8))/100.0; o+=2
            for i in range(5):
                r=data[o]|(data[o+1]<<8)
                if r&0x8000: r=-((~r+1)&0xFFFF)
                self.BattTemperatures[i]=r/100.0; o+=2
            for i in range(5):
                r=data[o]|(data[o+1]<<8)
                if r&0x8000: r=-((~r+1)&0xFFFF)
                self.PCBTemperatures[i]=r/100.0; o+=2
            r=data[o]|(data[o+1]<<8)
            if r&0x8000: r=-((~r+1)&0xFFFF)
            self.MinBattTemp=r/100.0; o+=2
            r=data[o]|(data[o+1]<<8)
            if r&0x8000: r=-((~r+1)&0xFFFF)
            self.MaxBattTemp=r/100.0; o+=2
            self.CellBalancingStatus=data[o]|(data[o+1]<<8); o+=2
            r=data[o]|(data[o+1]<<8)
            if r&0x8000: r=-((~r+1)&0xFFFF)
            self.CellTempDiff=r/1000.0; o+=2
            self.StackVoltage=(data[o]|(data[o+1]<<8))/100.0; o+=2
            self.Reserved1=data[o]; self.Reserved2=data[o+1]; o+=2
            self.BMSOperationState=data[o]; o+=1
            self.FETStatus=data[o]; self.FET_Status=self.FETStatus; o+=1
            self.SOC=data[o]; o+=1; self.SOH=data[o]; o+=1
            self.SOP=data[o]; o+=1; self.SOCDiff=data[o]; o+=1
            self.BinaryFlags=data[o]; o+=1; self.Reserved3=data[o]; o+=1
            self.Fault1=data[o]; o+=1; self.Fault2=data[o]; o+=1
            self.ThermalReason=data[o]; o+=1; self.Fault3=data[o]; o+=1
            self.ThermalRunawayFault=bool(data[o]); o+=1
            self.Reserved4=data[o]; o+=1
            self.parse_fault_bytes()
            return True
        except IndexError:
            print(f"❌ Payload too short at offset {o} (len={total})")
            return False
        except Exception as e:
            print(f"❌ Unexpected: {e}"); traceback.print_exc(); return False


# =============================================================================
# BLE MONITOR THREAD — INTEGRATED BMS + OTA (FULLY FIXED)
# =============================================================================

class BLEMonitorThread(threading.Thread):
    def __init__(self, gui_callback, bat_logger, sd_logger):
        super().__init__()
        self.gui_callback = gui_callback
        self.bat_logger   = bat_logger
        self.sd_logger    = sd_logger
        self.data         = BatteryData()
        self.daemon       = True
        self.running      = True
        self.loop         = None
        self.client       = None

        self.connect_requested     = False
        self.disconnect_requested  = False
        self.scan_requested        = False
        self.target_device_name    = DEFAULT_DEVICE_NAME
        self.target_device_address = None
        self.available_devices     = []
        self._lock                 = threading.Lock()

        # Battery streaming
        self.cmd_01_active   = False
        self.cmd_01_task     = None
        self.last_gui_update = 0

        # SD card
        self._sd_expected        = 0
        self._sd_received        = 0
        self._sd_records         = 0
        self._sd_log_dir         = None
        self._sd_active          = False
        self._sd_bat_was_01      = False
        self._sd_watchdog_handle = None
        self._sd_watchdog_armed  = False

        # BMS Config
        self._bms_cfg_ack_event   = threading.Event()
        self._bms_cfg_last_ack    = False
        self._bms_cfg_timeout     = BMS_CFG_ACK_TIMEOUT
        self._bms_cfg_max_retries = BMS_CFG_MAX_RETRIES
        self._bms_cfg_queue       = []
        self._bms_cfg_sending     = False
        self._bms_cfg_bat_was     = False
        self._bms_cfg_retry_map   = {}

        # OTA state - FIXED to match ota.py v3.0 exactly
        self.ota_requested   = False
        self.ota_filepath    = None
        self.ota_in_progress = False
        self._ota_ack_event  = None   # Created in run() on the correct loop
        self._ota_ack_ok     = False

    # ── battery streaming ─────────────────────────────────────────────

    def _start_bat_stream(self):
        self.cmd_01_active = True
        if self.loop and self.loop.is_running() and self.data.connected:
            if self.cmd_01_task and not self.cmd_01_task.done():
                self.cmd_01_task.cancel()
            self.cmd_01_task = asyncio.run_coroutine_threadsafe(
                self._cmd_01_loop(), self.loop)
        self.gui_callback(("streaming_status", "STREAMING"))
        self.gui_callback(("command_sent", "Battery streaming started (0x01)"))

    def _stop_bat_stream(self):
        self.cmd_01_active = False
        if self.cmd_01_task and not self.cmd_01_task.done():
            self.cmd_01_task.cancel()
        self.gui_callback(("streaming_status", "PAUSED"))

    async def _cmd_01_loop(self):
        while self.cmd_01_active and self.data.connected:
            await self._send_frame(build_frame(CMD_BAT_START))
            for _ in range(int(COMMAND_INTERVAL * 10)):
                if not self.cmd_01_active: break
                await asyncio.sleep(0.1)

    # ── SD watchdog ───────────────────────────────────────────────────

    def _arm_sd_watchdog(self):
        self._cancel_sd_watchdog()
        self._sd_watchdog_armed = True
        if self.loop and self.loop.is_running():
            self._sd_watchdog_handle = self.loop.call_later(
                SD_RESPONSE_TIMEOUT_SEC, self._sd_watchdog_fired)

    def _reset_sd_watchdog(self):
        # Cancel existing timer
        if self._sd_watchdog_handle is not None:
            try:
             self._sd_watchdog_handle.cancel()
            except Exception:
                pass
            self._sd_watchdog_handle = None
    
        # Only re-arm if SD read is still active
        if self._sd_active and self.loop and self.loop.is_running():
            self._sd_watchdog_handle = self.loop.call_later(
            SD_RESPONSE_TIMEOUT_SEC, self._sd_watchdog_fired)

    def _cancel_sd_watchdog(self):
        if self._sd_watchdog_handle is not None:
            try: self._sd_watchdog_handle.cancel()
            except Exception: pass
            self._sd_watchdog_handle = None
        self._sd_watchdog_armed = False

    def _sd_watchdog_fired(self):
        self._sd_watchdog_handle = None
        self._sd_watchdog_armed  = False
        if self.sd_logger.logging_active: self.sd_logger.stop_logging()
        self._sd_active   = False
        self._sd_expected = 0
        self._sd_received = 0
        self._sd_records  = 0
        self.gui_callback(("sd_timeout",
                           f"No SD response in {SD_RESPONSE_TIMEOUT_SEC:.0f}s"))
        self._resume_bat_stream(force=True)

    # ── SD card commands ──────────────────────────────────────────────

    def set_sd_log_directory(self, d):
        self._sd_log_dir = d

    def send_sd_info(self):
        payload = bytearray(9); payload[0] = SD_CARD_INFO
        asyncio.run_coroutine_threadsafe(
            self._send_frame(build_frame(CMD_SD_DATA, bytes(payload))), self.loop)
        self.gui_callback(("command_sent", "SD Card Info (0xAA) sent"))

    def send_sd_read(self, start_sector: int, end_sector: int):
        if not self._sd_log_dir:
            self.gui_callback(("sd_error", "No SD log directory selected")); return
        self._sd_bat_was_01 = self.cmd_01_active
        self._cancel_sd_watchdog()
        self._sd_active = False; self._sd_expected = 0
        self._sd_received = 0;   self._sd_records  = 0
        num_sectors       = end_sector - start_sector + 1
        self._sd_expected = num_sectors * SUBFRAMES_PER_SECTOR
        payload = bytearray(9); payload[0] = SD_READ_CARD
        for b in range(4):
            payload[1+b] = (start_sector >> (b*8)) & 0xFF
            payload[5+b] = (end_sector   >> (b*8)) & 0xFF
        asyncio.run_coroutine_threadsafe(
            self._send_frame(build_frame(CMD_SD_DATA, bytes(payload))), self.loop)
        if self.cmd_01_active: self._stop_bat_stream()
        ok, msg = self.sd_logger.start_logging(self._sd_log_dir)
        self.gui_callback(("sd_log_started", msg) if ok else ("sd_error", msg))
        self._sd_active = True
        self._arm_sd_watchdog()
        self.gui_callback(("sd_read_started",
                           (start_sector, end_sector,
                            self._sd_expected, self.sd_logger.file_path)))

    def send_sd_erase(self, start_sector: int, end_sector: int):
        self._sd_bat_was_01 = self.cmd_01_active
        if self.cmd_01_active: self._stop_bat_stream()
        payload = bytearray(9); payload[0] = SD_ERASE_CARD
        for b in range(4):
            payload[1+b] = (start_sector >> (b*8)) & 0xFF
            payload[5+b] = (end_sector   >> (b*8)) & 0xFF
        asyncio.run_coroutine_threadsafe(
            self._send_frame(build_frame(CMD_SD_DATA, bytes(payload))), self.loop)
        self.gui_callback(("command_sent",
                           f"SD Erase sent: sectors {start_sector}–{end_sector}"))
        self._arm_sd_watchdog()

    # ── BMS config ────────────────────────────────────────────────────

    def send_bms_config_all(self, param_list):
        if not self.data.connected:
            self.gui_callback(("bms_cfg_error", "Not connected")); return
        if self._bms_cfg_sending:
            self.gui_callback(("bms_cfg_error", "Send already in progress")); return
        self._bms_cfg_queue     = list(param_list)
        self._bms_cfg_sending   = True
        self._bms_cfg_bat_was   = self.cmd_01_active
        self._bms_cfg_retry_map = {}
        if self.cmd_01_active: self._stop_bat_stream()
        asyncio.run_coroutine_threadsafe(self._bms_cfg_send_loop(), self.loop)

    async def _bms_cfg_send_loop(self):
        total = len(self._bms_cfg_queue)
        success = 0; failed = []; timeout_params = []
        self.gui_callback(("bms_cfg_started", total))
        for idx, param in enumerate(self._bms_cfg_queue):
            param_name  = param['signal_name']
            retry_count = 0; param_success = False
            while retry_count < self._bms_cfg_max_retries and not param_success:
                try:
                    self._bms_cfg_ack_event.clear()
                    self._bms_cfg_last_ack = False
                    payload = build_bms_config_payload(
                        param['address'], param['data_type'], param['value'])
                    frame = build_frame(CMD_BMS_CONFIG, payload)
                    await self._send_frame(frame)
                    ack_received = await asyncio.wait_for(
                        self._wait_for_bms_ack(), timeout=self._bms_cfg_timeout)
                    if ack_received:
                        param_success = True; success += 1
                        self.gui_callback(("bms_cfg_progress",
                                         (idx+1, total, param_name, True, retry_count+1)))
                    else:
                        retry_count += 1
                        if retry_count < self._bms_cfg_max_retries:
                            self.gui_callback(("bms_cfg_retry",
                                             (param_name, retry_count, self._bms_cfg_max_retries)))
                            await asyncio.sleep(0.3)
                        else:
                            failed.append(param_name)
                            self.gui_callback(("bms_cfg_progress",
                                             (idx+1, total, param_name, False, retry_count)))
                except asyncio.TimeoutError:
                    retry_count += 1
                    if retry_count < self._bms_cfg_max_retries:
                        self.gui_callback(("bms_cfg_timeout_retry",
                                         (param_name, retry_count, self._bms_cfg_max_retries)))
                        await asyncio.sleep(0.5)
                    else:
                        timeout_params.append(param_name); failed.append(param_name)
                        self.gui_callback(("bms_cfg_progress",
                                         (idx+1, total, param_name, False, retry_count)))
                except Exception as e:
                    retry_count += 1
                    if retry_count >= self._bms_cfg_max_retries:
                        failed.append(param_name)
                        self.gui_callback(("bms_cfg_progress",
                                         (idx+1, total, param_name, False, retry_count)))
                if not param_success and retry_count < self._bms_cfg_max_retries:
                    await asyncio.sleep(0.2)
            await asyncio.sleep(0.1)
        self._bms_cfg_sending = False; self._bms_cfg_queue = []
        if timeout_params: self.gui_callback(("bms_cfg_timeout", timeout_params))
        self.gui_callback(("bms_cfg_done", (success, total, failed)))
        if self._bms_cfg_bat_was: self._start_bat_stream()

    async def _wait_for_bms_ack(self):
        await asyncio.get_event_loop().run_in_executor(
            None, self._bms_cfg_ack_event.wait, self._bms_cfg_timeout)
        if self._bms_cfg_ack_event.is_set():
            return self._bms_cfg_last_ack
        else:
            raise asyncio.TimeoutError()

    # ── OTA update — EXACT MATCH to ota.py v3.0 ───────────────────────

    def request_ota(self, filepath: str):
        if not self.data.connected:
            self.gui_callback(("ota_error", "Not connected")); return
        if self.ota_in_progress:
            self.gui_callback(("ota_error", "OTA already in progress")); return
        self.ota_filepath  = filepath
        self.ota_requested = True

    async def _run_ota_update(self, filepath: str) -> bool:
        """EXACT copy of ota.py v3.0 _run_ota_update"""
        try:
            with open(filepath, "rb") as f:
                firmware = f.read()
        except Exception as e:
            self.gui_callback(("ota_log", f"[OTA] Cannot open file: {e}"))
            return False

        file_size = len(firmware)
        remainder = file_size % CHUNK_SIZE
        if remainder:
            padded = firmware + bytes(CHUNK_SIZE - remainder)
        else:
            padded = firmware
        padded_size   = len(padded)
        total_chunks  = padded_size // CHUNK_SIZE
        crc_val       = crc32_firmware(padded)

        self.gui_callback(("ota_started", file_size, total_chunks))
        self.gui_callback(("ota_log",
            f"[OTA] File: {file_size}B  padded: {padded_size}B  "
            f"chunks: {total_chunks}  CRC: 0x{crc_val:08X}"))

        bat_was = self.cmd_01_active
        if bat_was: self._stop_bat_stream()

        # Step 1: TRIGGER
        self.gui_callback(("ota_step", "trigger", "active"))
        self.gui_callback(("ota_log", "[OTA] Step 1/5 — TRIGGER (0x36)"))
        if not await self._ota_send_subcmd(OTA_SUBCMD_TRIGGER, label="TRIGGER"):
            self.gui_callback(("ota_step", "trigger", "fail"))
            if bat_was: self._start_bat_stream()
            return False
        self.gui_callback(("ota_step", "trigger", "done"))

        # Step 2: FLASH_CMD
        self.gui_callback(("ota_step", "flash_cmd", "active"))
        self.gui_callback(("ota_log", "[OTA] Step 2/5 — FLASH_CMD (0x33) erasing..."))
        if not await self._ota_send_subcmd(OTA_SUBCMD_FLASH_CMD, label="FLASH_CMD"):
            self.gui_callback(("ota_step", "flash_cmd", "fail"))
            if bat_was: self._start_bat_stream()
            return False
        self.gui_callback(("ota_step", "flash_cmd", "done"))

        # Step 3: FILE_SIZE
        self.gui_callback(("ota_step", "file_size", "active"))
        size_bytes = struct.pack("<I", padded_size)
        self.gui_callback(("ota_log", f"[OTA] Step 3/5 — FILE_SIZE ({padded_size}B)"))
        if not await self._ota_send_subcmd(OTA_SUBCMD_FILE_SIZE,
                                            extra=size_bytes, label="FILE_SIZE"):
            self.gui_callback(("ota_step", "file_size", "fail"))
            if bat_was: self._start_bat_stream()
            return False
        self.gui_callback(("ota_step", "file_size", "done"))

        # Step 4: CHUNKS
        self.gui_callback(("ota_step", "chunks", "active"))
        for i in range(total_chunks):
            chunk = padded[i * CHUNK_SIZE:(i + 1) * CHUNK_SIZE]
            pct   = (i + 1) * 100 // total_chunks
            self.gui_callback(("ota_progress", i+1, total_chunks, pct))
            if not await self._ota_send_subcmd(OTA_SUBCMD_CHUNK, extra=chunk,
                                                label=f"CHUNK[{i+1}/{total_chunks}]"):
                self.gui_callback(("ota_step", "chunks", "fail"))
                if bat_was: self._start_bat_stream()
                return False
        self.gui_callback(("ota_step", "chunks", "done"))

        # Step 5: CRC
        self.gui_callback(("ota_step", "crc", "active"))
        crc_bytes = struct.pack("<I", crc_val)
        self.gui_callback(("ota_log", f"[OTA] Step 5/5 — CRC (0x{crc_val:08X})"))
        if not await self._ota_send_subcmd(OTA_SUBCMD_CRC,
                                            extra=crc_bytes, label="CRC"):
            self.gui_callback(("ota_step", "crc", "fail"))
            if bat_was: self._start_bat_stream()
            return False
        self.gui_callback(("ota_step", "crc", "done"))
        return True

    async def _ota_send_subcmd(self, subcmd: int,
                                extra: bytes = b"", label: str = "") -> bool:
        """EXACT copy of ota.py v3.0 _send_subcmd"""
        payload = bytes([subcmd]) + extra
        frame   = build_frame(CMD_OTA_UPDATE, payload)
        for attempt in range(1, OTA_MAX_RETRIES + 1):
            await self._send_frame(frame)
            ok = await self._ota_wait_ack()
            if ok:
                self.gui_callback(("ota_log", f"[{label}] ACK OK (attempt {attempt})"))
                return True
            self.gui_callback(("ota_log",
                f"[{label}] Attempt {attempt}/{OTA_MAX_RETRIES} FAIL — retry"))
            await asyncio.sleep(0.3)
        self.gui_callback(("ota_log", f"[{label}] FAILED after {OTA_MAX_RETRIES} attempts"))
        return False

    async def _ota_wait_ack(self) -> bool:
        """EXACT copy of ota.py v3.0 _wait_ack"""
        self._ota_ack_event.clear()
        try:
            await asyncio.wait_for(self._ota_ack_event.wait(), timeout=OTA_ACK_TIMEOUT)
            return self._ota_ack_ok
        except asyncio.TimeoutError:
            self.gui_callback(("ota_log", "[ACK] Timeout waiting for response"))
            return False

    # ── frame TX ─────────────────────────────────────────────────────

    async def _send_frame(self, frame):
        if self.client and self.client.is_connected:
            try:
                await self.client.write_gatt_char(REQUEST_UUID, bytes(frame), response=True)
            except Exception as e:
                print(f"  TX error: {e}")

    # ── public request methods ────────────────────────────────────────

    def request_scan(self):
        with self._lock: self.scan_requested = True

    def request_connect(self, name=None, address=None):
        with self._lock:
            self.connect_requested = True
            if name:    self.target_device_name    = name
            if address: self.target_device_address = address
            else:       self.target_device_address = None

    def request_disconnect(self):
        with self._lock: self.disconnect_requested = True

    def get_available_devices(self):
        with self._lock: return self.available_devices.copy()

    def stop(self):
        self.running = False
        if self.loop and self.loop.is_running():
            self.loop.call_soon_threadsafe(self.loop.stop)

    # ── thread main ───────────────────────────────────────────────────

    def run(self):
        self.loop = asyncio.new_event_loop()
        asyncio.set_event_loop(self.loop)
        # CRITICAL FIX: create asyncio.Event on the correct loop
        self._ota_ack_event = asyncio.Event()
        try:
            self.loop.run_until_complete(self._monitor_loop())
        except Exception as e:
            print(f"BLE thread error: {e}")
        finally:
            self.loop.close()

    async def _monitor_loop(self):
        # CRITICAL FIX: sleep 0.1s for snappy OTA pickup (matches ota.py)
        while self.running:
            with self._lock:
                scan_now  = self.scan_requested;  self.scan_requested = False
                conn_now  = self.connect_requested and not self.data.connected
                if conn_now: self.connect_requested = False
                disc_now  = self.disconnect_requested and self.data.connected
                if disc_now: self.disconnect_requested = False
                ota_now   = self.ota_requested and not self.ota_in_progress and self.data.connected
                if ota_now:
                    self.ota_requested   = False
                    self.ota_in_progress = True
                    fp = self.ota_filepath

            if scan_now:  await self._scan()
            if conn_now:  await self._connect()
            if disc_now:  await self._disconnect()
            if ota_now:
                success = await self._run_ota_update(fp)
                self.gui_callback(("ota_complete", success))
                self.ota_in_progress = False

            await asyncio.sleep(0.1)

    async def _scan(self):
        self.gui_callback(("scan_started", None))
        try:
            devices  = await BleakScanner.discover(timeout=5.0)
            dev_list = [{'name':d.name,'address':d.address,
                         'rssi':d.rssi if hasattr(d,'rssi') else 0}
                        for d in devices if d.name and d.name.strip()]
            dev_list.sort(key=lambda x: x['rssi'], reverse=True)
            with self._lock: self.available_devices = dev_list
            self.gui_callback(("scan_results", dev_list))
        except Exception as e:
            self.gui_callback(("scan_error", str(e)))

    async def _connect(self):
        self.gui_callback(("status", f"Connecting to {self.target_device_name}…"))
        try:
            if self.target_device_address:
                devs = await BleakScanner.discover(timeout=2.0)
                dev  = next((d for d in devs
                             if d.address.lower()==self.target_device_address.lower()), None)
            else:
                devs = await BleakScanner.discover(timeout=5.0)
                dev  = next((d for d in devs
                             if d.name and self.target_device_name.lower() in d.name.lower()), None)
            if not dev:
                self.data.connected = False
                self.gui_callback(self.data)
                self.gui_callback(("status","Device not found")); return
            self.client = BleakClient(dev, timeout=20.0)
            await self.client.connect()
            if self.client.is_connected:
                self.data.connected = True
                self.gui_callback(self.data)
                for service in self.client.services:
                    for char in service.characteristics:
                        if char.uuid.lower() == NOTIFY_UUID.lower():
                            if "notify" in char.properties:
                                await self.client.start_notify(
                                    char.uuid, self._notification_handler)
                self.gui_callback(("connected", None))
                self._start_bat_stream()
            else:
                self.data.connected = False; self.gui_callback(self.data)
        except Exception as e:
            print(f"Connect error: {e}")
            self.data.connected = False; self.gui_callback(self.data)

    async def _disconnect(self):
        self._cancel_sd_watchdog(); self._sd_active = False
        if self.client and self.client.is_connected:
            self.cmd_01_active = False
            if self.cmd_01_task and not self.cmd_01_task.done():
                self.cmd_01_task.cancel()
            try:
                await self.client.write_gatt_char(
                    REQUEST_UUID, build_frame(CMD_BAT_STOP), response=True)
                await asyncio.sleep(0.3)
            except Exception: pass
            try:
                await self.client.disconnect()
            except Exception: pass
            finally:
                self.client = None; self.data.connected = False
                self.gui_callback(self.data)

    def _notification_handler(self, characteristic, raw):
        """
        Handles all incoming BLE notifications.
        CRITICAL FIX for OTA: exactly matches ota.py v3.0
        """
        try:
            if len(raw) < 6 or raw[0] != SOF_BYTE_0 or raw[1] != SOF_BYTE_1: 
                return
            cmd         = raw[3]
            payload_len = (raw[4]<<8) | raw[5]
            if len(raw) < 6 + payload_len: 
                return
            payload = raw[6:6+payload_len]

            if cmd == CMD_BAT_START:
                self.data.notification_count += 1
                self.data.packet_count        = self.data.notification_count
                self.data.last_update         = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.data.update_from_payload(payload)
                if self.bat_logger.logging_active: 
                    self.bat_logger.log_data(self.data)
                now = datetime.now().timestamp()
                if now - self.last_gui_update > 0.1:
                    self.gui_callback(self.data)
                    self.last_gui_update = now

            elif cmd == CMD_BAT_STOP:
                self.gui_callback(("command_response","ACK 0x02"))

            elif cmd == CMD_SD_DATA:
                self._handle_sd_payload(payload)

            elif cmd == CMD_BMS_CONFIG:
                ack_ok = (len(payload) >= 1 and payload[0] == 0x01)
                self._bms_cfg_last_ack = ack_ok
                self._bms_cfg_ack_event.set()
                self.gui_callback(("bms_cfg_ack", ack_ok))

            elif cmd == CMD_OTA_UPDATE:
                # CRITICAL FIX: exactly mirrors ota.py v3.0 notification handler
                ack_ok = (len(payload) >= 1 and payload[0] == OTA_ACK_OK)
                self._ota_ack_ok = ack_ok
                self.gui_callback(("ota_ack_raw", ack_ok))
                if self.loop and self.loop.is_running():
                    asyncio.run_coroutine_threadsafe(
                        self._set_ota_ack_event(), self.loop)

        except Exception as e:
            print(f"Notification error: {e}")

    async def _set_ota_ack_event(self):
        """Set the OTA ack event on the asyncio loop — identical to ota.py."""
        self._ota_ack_event.set()

    def _handle_sd_payload(self, payload):
        if len(payload) < 1: return
        if len(payload) == 128 and payload[0] == 0xAA and payload[1] == 0x55:
            bms_bytes = bytes(payload[2:126])
            self._sd_received += 1; self._reset_sd_watchdog()
            signals = BMSSignalDecoder.decode(bms_bytes)
            if signals:
                self._sd_records += 1
                self._reset_sd_watchdog()
                ok = self.sd_logger.log_record(signals, self._sd_records)
                self.gui_callback(("sd_subframe",
                                   (self._sd_received, self._sd_expected,
                                    self._sd_records, signals,
                                    self.sd_logger.file_path, ok)))
            else:
                self.gui_callback(("sd_decode_error", self._sd_received))
            if (self._sd_active and self._sd_expected > 0
                    and self._sd_received >= self._sd_expected):
                self._cancel_sd_watchdog() 
                self._finish_sd_read()
            return
        if len(payload) == 19 and not (payload[0]==0xAA and payload[1]==0x55):
            info = {
                'card_type'        : payload[0],
                'capacity_gb'      : payload[1],
                'total_sectors'    : int.from_bytes(payload[2:6],   'little'),
                'manufacture_year' : int.from_bytes(payload[6:8],   'little'),
                'remaining_sectors': int.from_bytes(payload[8:12],  'little'),
                'manufacture_month': payload[12],
                'manufacturer_id'  : payload[13],
                'product_version'  : payload[14],
                'current_sector'   : int.from_bytes(payload[15:19], 'little'),
            }
            self.gui_callback(("sd_card_info", info)); return
        if len(payload) == 10 and payload[0] == SD_ERASE_CARD:
            self._cancel_sd_watchdog()
            success   = (payload[1] == 0x01)
            start_sec = int.from_bytes(payload[2:6],  'little')
            end_sec   = int.from_bytes(payload[6:10], 'little')
            self.gui_callback(("sd_erase_result", (success, start_sec, end_sec)))
            self._resume_bat_stream(); return

    def _finish_sd_read(self):
        self._cancel_sd_watchdog(); self._sd_active = False
        ok, msg = self.sd_logger.stop_logging()
        self.gui_callback(("sd_read_complete",
                           (self._sd_records, msg, self.sd_logger.file_path)))
        self._resume_bat_stream()

    def _resume_bat_stream(self, force: bool = False):
        if force or self._sd_bat_was_01: self._start_bat_stream()
        self._sd_bat_was_01 = False


# =============================================================================
# MAIN GUI — v5.2 (OTA FULLY FIXED)
# =============================================================================

class VPUSH_BMS_GUI:
    # Colour palette
    bg_dark        = '#060616'
    bg_card        = '#0d1229'
    bg_card2       = '#111830'
    accent_blue    = '#00d4ff'
    accent_green   = '#00ff95'
    accent_red     = '#ff4d4d'
    accent_orange  = '#ffaa00'
    accent_purple  = '#9d4edd'
    accent_yellow  = '#ffd966'
    accent_cyan    = '#00e5cc'
    text_primary   = '#e8f0fe'
    text_secondary = '#6b7fa3'
    temp_warning   = 45
    temp_critical  = 60

    font_title        = ('Consolas', 22, 'bold')
    font_header       = ('Consolas', 12, 'bold')
    font_normal       = ('Consolas', 10)
    font_small        = ('Consolas', 9)
    font_value_large  = ('Consolas', 34, 'bold')
    font_value_medium = ('Consolas', 18, 'bold')
    font_mono         = ('Courier New', 9)

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("⚡ VPUSH BMS DASHBOARD v5.2 — INTEGRATED OTA (FULLY FIXED)")
        self.root.geometry("1500x960")
        self.root.configure(bg=self.bg_dark)
        self.root.minsize(1200, 700)

        self.bat_logger = BatteryDataLogger()
        self.sd_logger  = SDExcelLogger()
        self.ble        = BLEMonitorThread(self._on_ble_event,
                                           self.bat_logger, self.sd_logger)
        self.ble.start()

        self._bms_params         = []
        self._bms_row_widgets    = []
        self._bms_cfg_excel_path = tk.StringVar(value="")
        self._bms_pending_count  = 0
        self._bms_ack_received   = 0
        self._bms_ack_ok         = 0
        self._bms_ack_fail       = 0

        # OTA step tracking
        self._ota_step_state = {k: 'idle' for k in
                                 ['trigger','flash_cmd','file_size','chunks','crc']}
        self._ota_total_chunks = 0
        self._ota_ack_count = 0  # FIXED: initialized in __init__

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ==================================================================
    # UI BUILD
    # ==================================================================

    def _build_ui(self):
        # Top header bar
        hdr = tk.Frame(self.root, bg=self.bg_card, height=64)
        hdr.pack(fill=tk.X, padx=0, pady=0)
        hdr.pack_propagate(False)

        bf = tk.Frame(hdr, bg=self.bg_card); bf.pack(side=tk.LEFT, padx=20, pady=12)
        tk.Label(bf, text="⚡", font=('Consolas',26), bg=self.bg_card,
                 fg=self.accent_blue).pack(side=tk.LEFT)
        tk.Label(bf, text=" VPUSH", font=self.font_title, bg=self.bg_card,
                 fg=self.text_primary).pack(side=tk.LEFT)
        tk.Label(bf, text=" BMS", font=self.font_title, bg=self.bg_card,
                 fg=self.accent_blue).pack(side=tk.LEFT)
        tk.Label(bf, text="  v5.2 — OTA FULLY FIXED", font=('Consolas',10),
                 bg=self.bg_card, fg=self.text_secondary).pack(side=tk.LEFT, padx=(4,0))

        sf = tk.Frame(hdr, bg=self.bg_card); sf.pack(side=tk.RIGHT, padx=20, pady=8)
        self.conn_status = tk.Label(sf, text="● DISCONNECTED",
                                    font=('Consolas',10,'bold'), bg=self.bg_card,
                                    fg=self.accent_red)
        self.conn_status.pack(anchor=tk.E)
        self.stream_status = tk.Label(sf, text="▶ STREAMING: OFF",
                                      font=self.font_small, bg=self.bg_card,
                                      fg=self.text_secondary)
        self.stream_status.pack(anchor=tk.E)
        self._ts_label = tk.Label(sf, text="", font=self.font_small,
                                  bg=self.bg_card, fg=self.text_secondary)
        self._ts_label.pack(anchor=tk.E)
        self._tick()

        nb = ttk.Notebook(self.root)
        nb.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        style = ttk.Style(); style.theme_use('clam')
        style.configure('TNotebook', background=self.bg_dark, borderwidth=0)
        style.configure('TNotebook.Tab', background='#111830',
                        foreground=self.text_secondary, padding=[18,6],
                        font=self.font_small)
        style.map('TNotebook.Tab',
                  background=[('selected','#00d4ff')],
                  foreground=[('selected','#000000')])
        style.configure('TProgressbar', troughcolor='#1a2240',
                        background='#00d4ff', thickness=18)

        tabs = [
            ("🔌  CONNECTION",  self._build_connection_tab),
            ("📊  DASHBOARD",   self._build_dashboard_tab),
            ("💾  SD CARD",     self._build_sd_tab),
            ("⚙️   BMS CONFIG",  self._build_bms_config_tab),
            ("🚀  OTA UPDATE",  self._build_ota_tab),
            ("📨  RAW DATA",    self._build_raw_tab),
        ]
        for title, builder in tabs:
            f = tk.Frame(nb, bg=self.bg_dark)
            nb.add(f, text=title)
            builder(f)

    def _tick(self):
        self._ts_label.config(text=datetime.now().strftime("%H:%M:%S"))
        self.root.after(1000, self._tick)

    # ------------------------------------------------------------------
    # CONNECTION TAB
    # ------------------------------------------------------------------

    def _build_connection_tab(self, parent):
        c = tk.Frame(parent, bg=self.bg_dark)
        c.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        sc = self._card(c, "⚙  DEVICE SETTINGS", self.accent_blue)
        sc.pack(fill=tk.X, pady=(0,12))
        sg = tk.Frame(sc, bg=self.bg_card); sg.pack(fill=tk.X, padx=16, pady=(0,16))
        self.device_name_var = tk.StringVar(value=DEFAULT_DEVICE_NAME)
        self.device_addr_var = tk.StringVar(value="")
        for lbl, var, hint in [("Device Name:", self.device_name_var, None),
                                ("Device Address:", self.device_addr_var, "(optional)")]:
            row = tk.Frame(sg, bg=self.bg_card); row.pack(fill=tk.X, pady=4)
            tk.Label(row, text=lbl, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary, width=18, anchor=tk.W).pack(side=tk.LEFT)
            tk.Entry(row, textvariable=var, font=self.font_normal,
                     bg='#0a1628', fg=self.text_primary, width=32,
                     insertbackground=self.accent_blue, relief='flat',
                     highlightthickness=1, highlightcolor=self.accent_blue,
                     highlightbackground='#1e2d50').pack(side=tk.LEFT, padx=8)
            if hint:
                tk.Label(row, text=hint, font=self.font_small,
                         bg=self.bg_card, fg=self.text_secondary).pack(side=tk.LEFT)
        for lbl, val, col in [
            ("Notify UUID:",  NOTIFY_UUID,  self.accent_blue),
            ("Request UUID:", REQUEST_UUID, self.accent_purple),
            ("OTA CMD:",      "0x05 — TRIGGER(0x36)→FLASH(0x33)→SIZE(0x01)→CHUNK(0x02)×N→CRC(0x03)", self.accent_orange),
            ("Frame format:", "[0x56][0x50][DIR][CMD][LEN_H][LEN_L][PAYLOAD][CS][0xAA][0xAA]", self.accent_yellow),
        ]:
            row = tk.Frame(sg, bg=self.bg_card); row.pack(fill=tk.X, pady=2)
            tk.Label(row, text=lbl, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary, width=18, anchor=tk.W).pack(side=tk.LEFT)
            tk.Label(row, text=val, font=('Courier New',9),
                     bg=self.bg_card, fg=col).pack(side=tk.LEFT, padx=8)

        cc = self._card(c, "🎮  CONNECTION CONTROL", self.accent_blue)
        cc.pack(fill=tk.X, pady=(0,12))
        bf2 = tk.Frame(cc, bg=self.bg_card); bf2.pack(fill=tk.X, padx=16, pady=(0,16))
        self.scan_btn = self._btn(bf2, "🔍  SCAN", self.accent_blue, 'black', self._scan)
        self.scan_btn.pack(side=tk.LEFT, padx=4)
        self.connect_btn = self._btn(bf2, "🔌  CONNECT", self.accent_green, 'black', self._connect)
        self.connect_btn.pack(side=tk.LEFT, padx=4)
        self.disconnect_btn = self._btn(bf2, "⏹  DISCONNECT", self.accent_red, 'white',
                                         self._disconnect, state=tk.DISABLED)
        self.disconnect_btn.pack(side=tk.LEFT, padx=4)
        self._btn(bf2, "🔄  REFRESH", self.accent_purple, 'white',
                  self._refresh_list).pack(side=tk.LEFT, padx=4)
        self.scan_status = tk.Label(cc, text="", font=self.font_small,
                                    bg=self.bg_card, fg=self.text_secondary)
        self.scan_status.pack(anchor=tk.W, padx=16, pady=(0,10))

        lc = self._card(c, "📋  AVAILABLE DEVICES", self.accent_blue)
        lc.pack(fill=tk.BOTH, expand=True)
        tf = tk.Frame(lc, bg=self.bg_card); tf.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0,16))
        sy = tk.Scrollbar(tf); sy.pack(side=tk.RIGHT, fill=tk.Y)
        self.device_tree = ttk.Treeview(tf, columns=('Name','Address','RSSI'),
                                        show='tree headings', yscrollcommand=sy.set, height=10)
        for col, w in [('#0',40),('Name',320),('Address',200),('RSSI',100)]:
            self.device_tree.column(col, width=w,
                                    anchor=tk.CENTER if col in ('#0','RSSI') else tk.W)
        for col, txt in [('#0','#'),('Name','Device Name'),
                          ('Address','BLE Address'),('RSSI','RSSI (dBm)')]:
            self.device_tree.heading(col, text=txt)
        self.device_tree.pack(fill=tk.BOTH, expand=True)
        sy.config(command=self.device_tree.yview)
        self.device_tree.bind('<Double-1>', lambda e: self._connect())
        s2 = ttk.Style()
        s2.configure("Treeview", background="#0a1628", foreground=self.text_primary,
                     fieldbackground="#0a1628", font=self.font_small, rowheight=24)
        s2.configure("Treeview.Heading", background=self.bg_card2,
                     foreground=self.accent_blue, font=self.font_small)

    # ------------------------------------------------------------------
    # DASHBOARD TAB
    # ------------------------------------------------------------------

    def _build_dashboard_tab(self, parent):
        dash_nb = ttk.Notebook(parent)
        dash_nb.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        for title, builder in [
            ("📊  LIVE DATA",      self._build_live_data_tab),
            ("⚠️   FAULTS & ALARMS", self._build_faults_tab),
            ("🔧  SYSTEM STATUS",  self._build_status_tab),
            ("📝  LIVE LOGGING",   self._build_logging_tab),
            ("📋  ACTIVITY LOG",   self._build_activity_tab),
        ]:
            f = tk.Frame(dash_nb, bg=self.bg_dark)
            dash_nb.add(f, text=title)
            builder(f)

    def _build_live_data_tab(self, parent):
        lp = tk.Frame(parent, bg=self.bg_dark, width=480)
        lp.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8,4), pady=8)
        lp.pack_propagate(False)
        rp = tk.Frame(parent, bg=self.bg_dark)
        rp.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(4,8), pady=8)
        self._build_overview_card(lp)
        self._build_fet_card(lp)
        self._build_capacity_card(lp)
        self._build_metrics_row(rp)
        self._build_temp_card(rp)
        self._build_stack_card(rp)
        self._build_bottom_bar(rp)
        bf = tk.Frame(parent, bg=self.bg_dark)
        bf.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True, padx=8, pady=(0,8))
        self._build_cells_card(bf)

    def _build_overview_card(self, p):
        card = self._card(p, "📊  BATTERY OVERVIEW", self.accent_blue)
        card.pack(fill=tk.X, pady=(0,8))
        c = tk.Frame(card, bg=self.bg_card); c.pack(fill=tk.X, padx=16, pady=(0,16))
        tk.Label(c, text="State of Charge", font=self.font_small,
                 bg=self.bg_card, fg=self.text_secondary).pack(anchor=tk.W)
        vf = tk.Frame(c, bg=self.bg_card); vf.pack(fill=tk.X)
        self.soc_label = tk.Label(vf, text="0", font=self.font_value_large,
                                  bg=self.bg_card, fg=self.accent_green)
        self.soc_label.pack(side=tk.LEFT)
        tk.Label(vf, text="%", font=('Consolas',16), bg=self.bg_card,
                 fg=self.text_secondary).pack(side=tk.LEFT, padx=(4,0), anchor=tk.S, pady=(0,6))
        bar_frame = tk.Frame(c, bg=self.bg_card); bar_frame.pack(fill=tk.X, pady=(4,12))
        self._soc_canvas = tk.Canvas(bar_frame, height=10, bg='#1a2240',
                                     highlightthickness=0)
        self._soc_canvas.pack(fill=tk.X)
        self._soc_bar_rect = self._soc_canvas.create_rectangle(
            0, 0, 0, 10, fill=self.accent_green, outline='')
        self._soc_canvas.bind("<Configure>", self._on_soc_resize)
        self._soc_pct = 0
        sf = tk.Frame(c, bg=self.bg_card); sf.pack(fill=tk.X)
        for attr, lbl, col in [('soh_label','SOH',self.accent_green),
                                ('sop_label','SOP',self.accent_blue)]:
            f = tk.Frame(sf, bg=self.bg_card); f.pack(side=tk.LEFT, expand=True)
            tk.Label(f, text=lbl, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary).pack()
            lw = tk.Label(f, text="0%", font=self.font_value_medium, bg=self.bg_card, fg=col)
            lw.pack(); setattr(self, attr, lw)

    def _on_soc_resize(self, event):
        self._draw_soc_bar(self._soc_pct)

    def _draw_soc_bar(self, pct: int):
        self._soc_pct = max(0, min(100, pct))
        w = self._soc_canvas.winfo_width()
        if w < 2: return
        fill_w = int(w * self._soc_pct / 100)
        color = (self.accent_red if pct < 15 else
                 self.accent_orange if pct < 30 else self.accent_green)
        self._soc_canvas.itemconfig(self._soc_bar_rect, fill=color)
        self._soc_canvas.coords(self._soc_bar_rect, 0, 0, fill_w, 10)

    def _build_fet_card(self, p):
        card = self._card(p, "🔌  FET STATUS", self.accent_cyan)
        card.pack(fill=tk.X, pady=(0,8))
        c = tk.Frame(card, bg=self.bg_card); c.pack(fill=tk.X, padx=16, pady=(0,16))
        for lbl, attr, init in [("Operation State","state_label","—"),
                                 ("FET Register","fet_label","0x00")]:
            f = tk.Frame(c, bg=self.bg_card); f.pack(fill=tk.X, pady=4)
            tk.Label(f, text=lbl, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary).pack(side=tk.LEFT)
            w = tk.Label(f, text=init, font=self.font_normal,
                         bg=self.bg_card, fg=self.accent_cyan)
            w.pack(side=tk.RIGHT); setattr(self, attr, w)
        ff = tk.Frame(c, bg=self.bg_card); ff.pack(fill=tk.X, pady=6)
        for i, (text, attr) in enumerate([("CHG", "chg_flag"),
                                           ("KEY", "key_flag"),
                                           ("BCU", "comm_flag")]):
            f = tk.Frame(ff, bg=self.bg_card); f.grid(row=0, column=i, padx=16)
            ind = tk.Label(f, text="●", font=('Consolas',14), bg=self.bg_card,
                           fg=self.text_secondary); ind.pack()
            tk.Label(f, text=text, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary).pack()
            setattr(self, attr, ind)

    def _build_capacity_card(self, p):
        card = self._card(p, "🔋  CAPACITY", self.accent_purple)
        card.pack(fill=tk.X, pady=(0,8))
        c = tk.Frame(card, bg=self.bg_card); c.pack(fill=tk.X, padx=16, pady=(0,16))
        for lbl, attr, init, col in [
            ("Charge Capacity",    "chg_cap_label",  "0.00 Ah", self.accent_green),
            ("Discharge Capacity", "dchg_cap_label", "0.00 Ah", self.accent_orange),
            ("SOC Difference",     "socdiff_label",  "0%",      self.text_primary),
        ]:
            f = tk.Frame(c, bg=self.bg_card); f.pack(fill=tk.X, pady=4)
            tk.Label(f, text=lbl, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary).pack(side=tk.LEFT)
            w = tk.Label(f, text=init, font=self.font_normal, bg=self.bg_card, fg=col)
            w.pack(side=tk.RIGHT); setattr(self, attr, w)

    def _build_metrics_row(self, p):
        mf = tk.Frame(p, bg=self.bg_dark); mf.pack(fill=tk.X, pady=(0,8))
        metrics = [
            ("Pack Voltage", "pack_volt_label", "0.00 V", self.accent_blue),
            ("Pack Current", "pack_curr_label", "0.00 A", self.accent_cyan),
            ("After FET",    "fet_volt_label",  "0.00 V", self.accent_purple),
            ("Load Voltage", "load_volt_label", "0.00 V", self.accent_orange),
        ]
        for lbl, attr, init, col in metrics:
            cc = tk.Frame(mf, bg=self.bg_card2, relief='flat')
            cc.pack(side=tk.LEFT, padx=3, expand=True, fill=tk.BOTH)
            tk.Label(cc, text=lbl, font=self.font_small, bg=self.bg_card2,
                     fg=self.text_secondary).pack(pady=(10,2))
            w = tk.Label(cc, text=init, font=self.font_value_medium,
                         bg=self.bg_card2, fg=col)
            w.pack(pady=(0,10)); setattr(self, attr, w)

    def _build_temp_card(self, p):
        card = self._card(p, "🌡️  TEMPERATURE", self.accent_orange)
        card.pack(fill=tk.X, pady=(0,8))
        hf = tk.Frame(card, bg=self.bg_card); hf.pack(fill=tk.X, padx=16, pady=(0,4))
        sf = tk.Frame(hf, bg=self.bg_card); sf.pack(side=tk.RIGHT)
        for attr, txt, col in [('max_temp_badge','MAX',self.accent_red),
                                ('min_temp_badge','MIN',self.accent_blue),
                                ('delta_temp_badge','ΔT',self.accent_yellow)]:
            mf2 = tk.Frame(sf, bg=self.bg_card); mf2.pack(side=tk.LEFT, padx=8)
            tk.Label(mf2, text=txt, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary).pack()
            w = tk.Label(mf2, text="0°C", font=self.font_normal, bg=self.bg_card, fg=col)
            w.pack(); setattr(self, attr, w)
        c = tk.Frame(card, bg=self.bg_card); c.pack(fill=tk.X, padx=16, pady=(0,12))
        tk.Label(c, text="Battery Sensors", font=self.font_small, bg=self.bg_card,
                 fg=self.accent_blue).pack(anchor=tk.W, pady=(4,4))
        btf = tk.Frame(c, bg=self.bg_card); btf.pack(fill=tk.X)
        self.batt_temp_labels = []
        for i in range(5):
            sf2 = tk.Frame(btf, bg='#0a1628', height=60)
            sf2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
            sf2.pack_propagate(False)
            tk.Label(sf2, text=f"TS{i+1}", font=self.font_small, bg='#0a1628',
                     fg=self.text_secondary).pack(pady=(8,2))
            lbl = tk.Label(sf2, text="0°C", font=self.font_normal,
                           bg='#0a1628', fg=self.text_primary); lbl.pack()
            self.batt_temp_labels.append(lbl)
        tk.Label(c, text="PCB Sensors", font=self.font_small, bg=self.bg_card,
                 fg=self.accent_purple).pack(anchor=tk.W, pady=(8,4))
        ptf = tk.Frame(c, bg=self.bg_card); ptf.pack(fill=tk.X)
        self.pcb_temp_labels = []
        for i in range(5):
            sf3 = tk.Frame(ptf, bg='#0d1030', height=60)
            sf3.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
            sf3.pack_propagate(False)
            tk.Label(sf3, text=f"P{i+1}", font=self.font_small, bg='#0d1030',
                     fg=self.text_secondary).pack(pady=(8,2))
            lbl = tk.Label(sf3, text="0°C", font=self.font_normal,
                           bg='#0d1030', fg=self.accent_purple); lbl.pack()
            self.pcb_temp_labels.append(lbl)
        stf = tk.Frame(c, bg=self.bg_card); stf.pack(fill=tk.X, pady=(8,0))
        self.temp_status = tk.Label(stf, text="✓ Temperature Normal",
                                    font=self.font_small, bg=self.bg_card,
                                    fg=self.accent_green)
        self.temp_status.pack(side=tk.LEFT)
        self.temp_warn_lbl = tk.Label(stf, text="", font=self.font_small,
                                      bg=self.bg_card, fg=self.accent_orange)
        self.temp_warn_lbl.pack(side=tk.RIGHT)

    def _build_stack_card(self, p):
        card = self._card(p, "🔋  STACK VOLTAGE", self.accent_purple)
        card.pack(fill=tk.X, pady=(0,8))
        f = tk.Frame(card, bg=self.bg_card); f.pack(fill=tk.X, padx=16, pady=(0,14))
        tk.Label(f, text="Stack Voltage:", font=self.font_small, bg=self.bg_card,
                 fg=self.text_secondary).pack(side=tk.LEFT)
        self.stack_volt_label = tk.Label(f, text="0.000 V", font=self.font_value_medium,
                                         bg=self.bg_card, fg=self.accent_purple)
        self.stack_volt_label.pack(side=tk.RIGHT)

    def _build_bottom_bar(self, p):
        b = tk.Frame(p, bg=self.bg_card2, height=30); b.pack(fill=tk.X)
        b.pack_propagate(False)
        inf = tk.Frame(b, bg=self.bg_card2); inf.pack(fill=tk.BOTH, padx=12)
        tk.Label(inf, text="BMS: VP_01", font=self.font_small,
                 bg=self.bg_card2, fg=self.accent_blue).pack(side=tk.LEFT)
        self.notif_count_label = tk.Label(inf, text="Pkt: 0",
                                          font=self.font_small, bg=self.bg_card2,
                                          fg=self.text_secondary)
        self.notif_count_label.pack(side=tk.RIGHT, padx=(0,16))
        self.timestamp_label = tk.Label(inf, text="Updated: Never",
                                        font=self.font_small, bg=self.bg_card2,
                                        fg=self.text_secondary)
        self.timestamp_label.pack(side=tk.RIGHT)

    def _build_cells_card(self, p):
        card = self._card(p, "🔋  CELL VOLTAGES  (16 cells)", self.accent_blue)
        card.pack(fill=tk.BOTH, expand=True)
        hf = tk.Frame(card, bg=self.bg_card); hf.pack(fill=tk.X, padx=16, pady=(0,4))
        sf = tk.Frame(hf, bg=self.bg_card); sf.pack(side=tk.RIGHT)
        self.max_cell_label  = self._stat_badge(sf, "Max", "0.000V", self.accent_blue)
        self.min_cell_label  = self._stat_badge(sf, "Min", "0.000V", self.accent_orange)
        self.diff_cell_label = self._stat_badge(sf, "Δ",   "0 mV",  self.accent_red)
        gf = tk.Frame(card, bg=self.bg_card)
        gf.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0,12))
        self.cell_labels = []
        for row in range(4):
            rf = tk.Frame(gf, bg=self.bg_card); rf.pack(fill=tk.X, expand=True, pady=2)
            for col in range(4):
                cn = row*4+col
                cc = tk.Frame(rf, bg='#0a1628', height=52)
                cc.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
                cc.pack_propagate(False)
                tk.Label(cc, text=f"C{cn+1}", font=self.font_small,
                         bg='#0a1628', fg=self.text_secondary).pack(pady=(6,0))
                lbl = tk.Label(cc, text="0.000V", font=self.font_normal,
                               bg='#0a1628', fg=self.text_primary); lbl.pack()
                self.cell_labels.append(lbl)

    def _stat_badge(self, p, lbl, val, col):
        f = tk.Frame(p, bg=self.bg_card); f.pack(side=tk.LEFT, padx=(12,0))
        tk.Label(f, text=lbl, font=self.font_small, bg=self.bg_card,
                 fg=self.text_secondary).pack(side=tk.LEFT, padx=(0,4))
        w = tk.Label(f, text=val, font=self.font_normal, bg=self.bg_card, fg=col)
        w.pack(side=tk.LEFT); return w

    def _build_faults_tab(self, parent):
        lf = tk.Frame(parent, bg=self.bg_dark)
        lf.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8,4), pady=8)
        rf = tk.Frame(parent, bg=self.bg_dark)
        rf.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(4,8), pady=8)
        self.fault_labels = {}
        self._fault_card(lf, "🔴  CELL FAULTS", [
            ("Min Cell UV Fault",'min_cell_uv'),("Max Cell OV Fault",'max_cell_ov'),
            ("Over Charge Current",'over_chg_curr'),("Over Discharge Current",'over_dchg_curr')])
        self._fault_card(lf, "🔴  BATTERY FAULTS", [
            ("Under Temperature",'batt_ut'),("Over Temperature",'batt_ot'),
            ("Over Voltage",'batt_ov'),("Under Voltage",'batt_uv')])
        self._thermal_card(rf)
        self._alarms_card(rf)

    def _fault_card(self, p, title, faults):
        card = self._card(p, title, self.accent_red)
        card.pack(fill=tk.X, pady=(0,10))
        c = tk.Frame(card, bg=self.bg_card); c.pack(fill=tk.X, padx=16, pady=(0,14))
        for label, key in faults:
            f = tk.Frame(c, bg=self.bg_card); f.pack(fill=tk.X, pady=3)
            tk.Label(f, text=label, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary).pack(side=tk.LEFT)
            lbl = tk.Label(f, text="OK", font=self.font_normal,
                           bg=self.bg_card, fg=self.accent_green); lbl.pack(side=tk.RIGHT)
            self.fault_labels[key] = lbl

    def _thermal_card(self, p):
        card = self._card(p, "🔥  THERMAL RUNAWAY", self.accent_orange)
        card.pack(fill=tk.X, pady=(0,10))
        hf = tk.Frame(card, bg=self.bg_card); hf.pack(fill=tk.X, padx=16, pady=(0,4))
        self.thermal_fault_label = tk.Label(hf, text="●", font=('Consolas',16),
                                            bg=self.bg_card, fg=self.accent_green)
        self.thermal_fault_label.pack(side=tk.RIGHT)
        c = tk.Frame(card, bg=self.bg_card); c.pack(fill=tk.X, padx=16, pady=(0,14))
        self.thermal_labels = {}
        for reason in ["Min Cell UV","Max Cell Temp","Max Cell Temp Rise",
                       "Cell Temp Diff","Max Cell OV","Cell Volt Sampling","Temp Sampling"]:
            f = tk.Frame(c, bg=self.bg_card); f.pack(fill=tk.X, pady=2)
            tk.Label(f, text=reason, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary).pack(side=tk.LEFT)
            lbl = tk.Label(f, text="✓", font=self.font_normal,
                           bg=self.bg_card, fg=self.accent_green); lbl.pack(side=tk.RIGHT)
            self.thermal_labels[reason] = lbl

    def _alarms_card(self, p):
        card = self._card(p, "⚠️   ALARMS", self.accent_yellow)
        card.pack(fill=tk.X, pady=(0,10))
        c = tk.Frame(card, bg=self.bg_card); c.pack(fill=tk.X, padx=16, pady=(0,14))
        self.alarm_labels = {}
        for alarm in ["Over PCB Temp","Pre-Charge Failure","Hardware SC",
                      "Cell ΔV Alarm","Cell ΔT Alarm","Low SOC Alarm",
                      "Low SOH Alarm","SOC Jump Alarm"]:
            f = tk.Frame(c, bg=self.bg_card); f.pack(fill=tk.X, pady=2)
            tk.Label(f, text=alarm, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary).pack(side=tk.LEFT)
            lbl = tk.Label(f, text="✓", font=self.font_normal,
                           bg=self.bg_card, fg=self.accent_green); lbl.pack(side=tk.RIGHT)
            self.alarm_labels[alarm] = lbl

    def _build_status_tab(self, parent):
        bc = self._card(parent, "⚖️   CELL BALANCING", self.accent_blue)
        bc.pack(fill=tk.X, padx=8, pady=(8,4))
        self.balance_label = tk.Label(bc, text="0x0000  (No cells balancing)",
                                      font=self.font_normal, bg=self.bg_card,
                                      fg=self.text_primary)
        self.balance_label.pack(padx=16, pady=(0,14))
        ic = self._card(parent, "ℹ️   SYSTEM INFO", self.accent_blue)
        ic.pack(fill=tk.X, padx=8, pady=(0,4))
        ig = tk.Frame(ic, bg=self.bg_card); ig.pack(fill=tk.X, padx=16, pady=(0,14))
        self.info_labels = {}
        for lbl, val in [("BMS Number","VP_01"),("Packet Count","0"),("Last Update","Never")]:
            f = tk.Frame(ig, bg=self.bg_card); f.pack(fill=tk.X, pady=2)
            tk.Label(f, text=lbl, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary).pack(side=tk.LEFT)
            w = tk.Label(f, text=val, font=self.font_normal, bg=self.bg_card,
                         fg=self.text_primary); w.pack(side=tk.RIGHT)
            self.info_labels[lbl] = w

    def _build_logging_tab(self, parent):
        c = tk.Frame(parent, bg=self.bg_dark)
        c.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        cc = self._card(c, "📝  LIVE BATTERY LOG CONTROL", self.accent_blue)
        cc.pack(fill=tk.X, pady=(0,12))
        bf = tk.Frame(cc, bg=self.bg_card); bf.pack(fill=tk.X, padx=16, pady=(0,16))
        self.start_log_btn = self._btn(bf, "▶  START LOGGING", self.accent_green, 'black',
                                        self._start_bat_log, w=18)
        self.start_log_btn.pack(side=tk.LEFT, padx=4)
        self.stop_log_btn = self._btn(bf, "⏹  STOP LOGGING", self.accent_red, 'white',
                                       self._stop_bat_log, w=18, state=tk.DISABLED)
        self.stop_log_btn.pack(side=tk.LEFT, padx=4)
        self._btn(bf, "📁  CHOOSE DIR", self.accent_blue, 'black',
                  self._choose_bat_dir, w=16).pack(side=tk.LEFT, padx=4)
        sc = self._card(c, "📊  STATUS", self.accent_blue)
        sc.pack(fill=tk.X, pady=(0,12))
        sg = tk.Frame(sc, bg=self.bg_card); sg.pack(fill=tk.X, padx=16, pady=(0,16))
        self.log_status_labels = {}
        for lbl, key, default in [
            ("Status:","log_status","Stopped"),("File:","log_file","Not selected"),
            ("Records:","log_count","0"),("Duration:","log_duration","—"),
            ("Directory:","log_dir","Not selected")]:
            f = tk.Frame(sg, bg=self.bg_card); f.pack(fill=tk.X, pady=3)
            tk.Label(f, text=lbl, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary, width=12, anchor=tk.W).pack(side=tk.LEFT)
            w = tk.Label(f, text=default, font=self.font_normal, bg=self.bg_card,
                         fg=self.text_primary); w.pack(side=tk.RIGHT, padx=8)
            self.log_status_labels[key] = w
        self.bat_log_dir = None

    def _build_activity_tab(self, parent):
        ldc = self._card(parent, "📋  ACTIVITY LOG", self.accent_blue)
        ldc.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        self.log_messages = scrolledtext.ScrolledText(
            ldc, height=20, bg='#080e22', fg='#00d4ff',
            font=self.font_mono, wrap=tk.WORD, insertbackground=self.accent_blue)
        self.log_messages.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0,16))

    # ------------------------------------------------------------------
    # SD CARD TAB
    # ------------------------------------------------------------------

    def _build_sd_tab(self, parent):
        c = tk.Frame(parent, bg=self.bg_dark)
        c.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        ctrl = self._card(c, "💾  SD CARD CONTROL", self.accent_blue)
        ctrl.pack(fill=tk.X, pady=(0,8))
        tk.Label(ctrl,
            text=f"ℹ  Streaming auto-pauses during SD operations and resumes on completion "
                 f"OR after {SD_RESPONSE_TIMEOUT_SEC:.0f}s timeout.",
            font=self.font_small, bg=self.bg_card, fg=self.accent_yellow,
            wraplength=900, justify=tk.LEFT).pack(anchor=tk.W, padx=16, pady=(0,8))
        bf = tk.Frame(ctrl, bg=self.bg_card); bf.pack(fill=tk.X, padx=16, pady=(0,16))
        self.sd_info_btn = self._btn(bf, "📋  Card Info", self.accent_blue, 'black',
                                      self._sd_info, state=tk.DISABLED, w=16)
        self.sd_info_btn.pack(side=tk.LEFT, padx=4)
        self.sd_read_btn = self._btn(bf, "📖  Read Sectors", self.accent_green, 'black',
                                      self._sd_read_dialog, state=tk.DISABLED, w=18)
        self.sd_read_btn.pack(side=tk.LEFT, padx=4)
        self.sd_erase_btn = self._btn(bf, "🗑  Erase", self.accent_orange, 'black',
                                       self._sd_erase_dialog, state=tk.DISABLED, w=14)
        self.sd_erase_btn.pack(side=tk.LEFT, padx=4)
        self._btn(bf, "📁  Select Dir", self.accent_purple, 'white',
                  self._sd_select_dir, w=14).pack(side=tk.LEFT, padx=4)
        sf = tk.Frame(ctrl, bg=self.bg_card); sf.pack(fill=tk.X, padx=16, pady=(0,4))
        sf2 = tk.Frame(sf, bg=self.bg_card); sf2.pack(fill=tk.X)
        self.sd_status_lbl = tk.Label(sf2, text="No SD operation yet",
                                      font=self.font_small, bg=self.bg_card,
                                      fg=self.accent_yellow)
        self.sd_status_lbl.pack(side=tk.LEFT)
        self.sd_log_indicator = tk.Label(sf2, text="⬤ IDLE", font=('Consolas',10,'bold'),
                                         bg=self.bg_card, fg='#333355')
        self.sd_log_indicator.pack(side=tk.RIGHT, padx=16)
        self.sd_counter_lbl = tk.Label(sf2, text="Records: 0 / 0",
                                       font=self.font_small, bg=self.bg_card,
                                       fg=self.text_primary)
        self.sd_counter_lbl.pack(side=tk.RIGHT)
        self.sd_dir_lbl = tk.Label(ctrl, text="No SD directory selected",
                                   font=self.font_small, bg=self.bg_card,
                                   fg=self.text_secondary)
        self.sd_dir_lbl.pack(anchor=tk.W, padx=16, pady=(2,10))
        sd_pb_frame = tk.Frame(ctrl, bg=self.bg_card)
        sd_pb_frame.pack(fill=tk.X, padx=16, pady=(0,12))
        self.sd_progress_bar = ttk.Progressbar(sd_pb_frame, length=300, mode='determinate')
        self.sd_progress_bar.pack(side=tk.LEFT)
        self.sd_progress_lbl = tk.Label(sd_pb_frame, text="", font=self.font_small,
                                         bg=self.bg_card, fg=self.text_secondary)
        self.sd_progress_lbl.pack(side=tk.LEFT, padx=8)
        tk.Label(c, text="📋  CARD INFO", font=self.font_header,
                 bg=self.bg_dark, fg=self.accent_blue).pack(anchor=tk.W)
        self.sd_info_text = tk.Text(c, height=5, bg='#080e22', fg=self.text_primary,
                                    font=self.font_mono, relief='flat')
        self.sd_info_text.pack(fill=tk.X, pady=(4,8))
        tk.Label(c, text="📊  SD DECODED SIGNALS", font=self.font_header,
                 bg=self.bg_dark, fg=self.accent_green).pack(anchor=tk.W)
        self.sd_data_text = scrolledtext.ScrolledText(
            c, height=14, font=self.font_mono, bg='#080e22', fg='#00ff9d')
        self.sd_data_text.pack(fill=tk.BOTH, expand=True)
        for tag, fg in [("header","#00d4ff"),("data","#00ff9d"),
                        ("done","#ffaa00"),("error","#ff4d4d"),("timeout","#ff9900")]:
            self.sd_data_text.tag_config(tag, foreground=fg,
                font=self.font_mono if tag not in ('header','done','timeout')
                else ('Courier New',9,'bold'))

    # ------------------------------------------------------------------
    # OTA TAB — v5.2, identical step flow to ota.py v3.0
    # ------------------------------------------------------------------

    def _build_ota_tab(self, parent):
        c = tk.Frame(parent, bg=self.bg_dark)
        c.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        # Protocol banner
        info = tk.Frame(c, bg='#0d1a30', relief='flat')
        info.pack(fill=tk.X, pady=(0,8))
        tk.Label(info, text="🚀  OTA FIRMWARE UPDATE — Fully automatic 5-step sequence",
                 font=self.font_header, bg='#0d1a30', fg=self.accent_blue).pack(
                     anchor=tk.W, padx=16, pady=(12,2))
        tk.Label(info, text=
                 "Frame: [0x56][0x50][DIR][CMD][LEN_H][LEN_L][PAYLOAD][CS][0xAA][0xAA]  "
                 "|  CS = ~(CMD+Σpayload)&0xFF  |  CRC-32 poly 0xEDB88320",
                 font=('Courier New',9), bg='#0d1a30', fg=self.accent_yellow).pack(
                     anchor=tk.W, padx=16, pady=(0,12))

        left = tk.Frame(c, bg=self.bg_dark); left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        right = tk.Frame(c, bg=self.bg_dark, width=380)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, padx=(8,0))
        right.pack_propagate(False)

        # Left: file + control + progress
        fc = self._card(left, "📁  FIRMWARE FILE", self.accent_blue)
        fc.pack(fill=tk.X, pady=(0,8))
        ff = tk.Frame(fc, bg=self.bg_card); ff.pack(fill=tk.X, padx=16, pady=(0,16))
        self.ota_file_var = tk.StringVar()
        tk.Entry(ff, textvariable=self.ota_file_var, font=self.font_normal,
                 bg='#0a1628', fg=self.text_primary, width=52,
                 insertbackground=self.accent_blue, relief='flat',
                 highlightthickness=1, highlightcolor=self.accent_blue,
                 highlightbackground='#1e2d50').pack(side=tk.LEFT, padx=(0,8))
        self._btn(ff, "📂  BROWSE", self.accent_blue, 'black',
                  self._ota_browse, w=12).pack(side=tk.LEFT)

        fi = self._card(left, "ℹ️   FILE INFORMATION", self.accent_blue)
        fi.pack(fill=tk.X, pady=(0,8))
        fig = tk.Frame(fi, bg=self.bg_card); fig.pack(fill=tk.X, padx=16, pady=(0,16))
        self.ota_file_info = {}
        for lbl in ["Path:", "Size:", "Padded Size:", "CRC32 (reflected):"]:
            f = tk.Frame(fig, bg=self.bg_card); f.pack(fill=tk.X, pady=3)
            tk.Label(f, text=lbl, font=self.font_small, bg=self.bg_card,
                     fg=self.text_secondary, width=20, anchor=tk.W).pack(side=tk.LEFT)
            w = tk.Label(f, text="—", font=('Courier New',9),
                         bg=self.bg_card, fg=self.text_primary)
            w.pack(side=tk.LEFT, padx=8); self.ota_file_info[lbl] = w

        oc = self._card(left, "⚡  OTA CONTROL", self.accent_blue)
        oc.pack(fill=tk.X, pady=(0,8))
        ob = tk.Frame(oc, bg=self.bg_card); ob.pack(fill=tk.X, padx=16, pady=(0,8))
        self.ota_start_btn = self._btn(ob, "⚡  START OTA UPDATE",
                                        self.accent_green, 'black',
                                        self._start_ota, w=22, state=tk.DISABLED)
        self.ota_start_btn.pack(side=tk.LEFT, padx=4)
        self.ota_status_lbl = tk.Label(ob, text="Select .bin file and connect to begin",
                                        font=self.font_small, bg=self.bg_card,
                                        fg=self.text_secondary)
        self.ota_status_lbl.pack(side=tk.LEFT, padx=12)

        pb_frame = tk.Frame(oc, bg=self.bg_card); pb_frame.pack(fill=tk.X, padx=16, pady=(0,4))
        tk.Label(pb_frame, text="Chunk Progress:", font=self.font_small,
                 bg=self.bg_card, fg=self.text_secondary).pack(anchor=tk.W)
        pb2 = tk.Frame(pb_frame, bg=self.bg_card); pb2.pack(fill=tk.X)
        self.ota_progress_bar = ttk.Progressbar(pb2, length=500, mode='determinate',
                                                  maximum=100, value=0)
        self.ota_progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=4)
        self.ota_progress_lbl = tk.Label(pb2, text="0 / 0", font=self.font_small,
                                          bg=self.bg_card, fg=self.text_secondary, width=12)
        self.ota_progress_lbl.pack(side=tk.LEFT, padx=6)
        self.ota_pct_lbl = tk.Label(oc, text="0%", font=('Consolas',11,'bold'),
                                     bg=self.bg_card, fg=self.accent_cyan)
        self.ota_pct_lbl.pack(anchor=tk.W, padx=16, pady=(0,12))

        log_card = self._card(left, "📨  OTA FRAME LOG", self.accent_blue)
        log_card.pack(fill=tk.BOTH, expand=True)
        self.ota_log_text = scrolledtext.ScrolledText(log_card, height=10,
                                                       font=self.font_mono,
                                                       bg='#080e22', fg='#00ff9d',
                                                       wrap=tk.WORD)
        self.ota_log_text.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0,16))
        for tag, fg in [("ok","#00ff95"),("err","#ff4d4d"),
                        ("info","#00d4ff"),("warning","#ffaa00")]:
            self.ota_log_text.tag_config(tag, foreground=fg)

        # Right: 5-step sequence indicator
        step_card = self._card(right, "📋  SEQUENCE STEPS", self.accent_blue)
        step_card.pack(fill=tk.X, pady=(0,8))
        sf2 = tk.Frame(step_card, bg=self.bg_card)
        sf2.pack(fill=tk.X, padx=16, pady=(0,16))
        self.ota_step_labels  = {}
        self.ota_step_percent = {}
        steps = [
            ("trigger",   "TRIGGER  (0x36)", "Prepare S32 / reset session"),
            ("flash_cmd", "FLASH    (0x33)", "Erase application flash"),
            ("file_size", "SIZE     (0x01)", "Send firmware image size"),
            ("chunks",    "CHUNKS   (0x02)", "Write firmware data blocks"),
            ("crc",       "CRC      (0x03)", "Verify integrity & reboot"),
        ]
        for i, (key, title, desc) in enumerate(steps):
            row = tk.Frame(sf2, bg='#0a1628'); row.pack(fill=tk.X, pady=3)
            tk.Label(row, text=f"{i+1}", font=('Consolas',11,'bold'), width=2,
                     bg='#0a1628', fg=self.text_secondary).pack(
                         side=tk.LEFT, padx=(8,6), pady=6)
            ind = tk.Label(row, text="○", font=('Consolas',14,'bold'), width=2,
                           bg='#0a1628', fg=self.text_secondary)
            ind.pack(side=tk.LEFT)
            detail = tk.Frame(row, bg='#0a1628'); detail.pack(side=tk.LEFT, padx=6, pady=4)
            tk.Label(detail, text=title, font=('Consolas',9,'bold'),
                     bg='#0a1628', fg=self.text_primary, anchor=tk.W).pack(anchor=tk.W)
            tk.Label(detail, text=desc, font=('Consolas',8),
                     bg='#0a1628', fg=self.text_secondary, anchor=tk.W).pack(anchor=tk.W)
            self.ota_step_labels[key] = ind

        chunk_detail = tk.Frame(step_card, bg=self.bg_card)
        chunk_detail.pack(fill=tk.X, padx=16, pady=(0,12))
        tk.Label(chunk_detail, text="Chunk detail:", font=self.font_small,
                 bg=self.bg_card, fg=self.text_secondary).pack(anchor=tk.W)
        self.ota_chunk_lbl = tk.Label(chunk_detail, text="—",
                                       font=self.font_small, bg=self.bg_card,
                                       fg=self.text_secondary)
        self.ota_chunk_lbl.pack(anchor=tk.W)

        ack_card = self._card(right, "✅  ACK HISTORY", self.accent_green)
        ack_card.pack(fill=tk.X, pady=(0,8))
        self.ota_ack_hist = scrolledtext.ScrolledText(
            ack_card, height=6, font=self.font_mono, bg='#080e22', fg='#00ff9d',
            wrap=tk.WORD, state=tk.NORMAL)
        self.ota_ack_hist.pack(fill=tk.X, padx=16, pady=(0,16))
        self.ota_ack_hist.tag_config("ok",  foreground="#00ff95")
        self.ota_ack_hist.tag_config("err", foreground="#ff4d4d")

    # ------------------------------------------------------------------
    # BMS CONFIG TAB
    # ------------------------------------------------------------------

    def _build_bms_config_tab(self, parent):
        c = tk.Frame(parent, bg=self.bg_dark)
        c.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        file_bar = self._card(c, "⚙️   BMS CONFIGURATION  (ACK-based pacing)", self.accent_blue)
        file_bar.pack(fill=tk.X, pady=(0,6))
        fb2 = tk.Frame(file_bar, bg=self.bg_card); fb2.pack(fill=tk.X, padx=16, pady=(0,12))
        tk.Label(fb2, text="Excel File:", font=self.font_small, bg=self.bg_card,
                 fg=self.text_secondary).pack(side=tk.LEFT)
        tk.Entry(fb2, textvariable=self._bms_cfg_excel_path, font=self.font_small,
                 bg='#0a1628', fg=self.text_primary, width=55, relief='flat',
                 highlightthickness=1, highlightcolor=self.accent_blue,
                 highlightbackground='#1e2d50').pack(side=tk.LEFT, padx=8)
        self._btn(fb2, "📂  Browse", self.accent_purple, 'white',
                  self._bms_cfg_browse, w=10).pack(side=tk.LEFT, padx=4)
        self.bms_cfg_load_btn = self._btn(fb2, "⬆  Load", self.accent_blue, 'black',
                                           self._bms_cfg_load, w=10)
        self.bms_cfg_load_btn.pack(side=tk.LEFT, padx=4)

        act_bar = tk.Frame(c, bg=self.bg_card); act_bar.pack(fill=tk.X, pady=(0,6))
        ab2 = tk.Frame(act_bar, bg=self.bg_card); ab2.pack(padx=16, pady=10)
        self.bms_send_all_btn = self._btn(ab2, "📡  SEND ALL", self.accent_green, 'black',
                                           self._bms_cfg_send_all, w=18, h=2, state=tk.DISABLED)
        self.bms_send_all_btn.pack(side=tk.LEFT, padx=4)
        self.bms_send_sel_btn = self._btn(ab2, "📡  SEND SELECTED", self.accent_orange, 'black',
                                           self._bms_cfg_send_selected, w=18, h=2, state=tk.DISABLED)
        self.bms_send_sel_btn.pack(side=tk.LEFT, padx=4)
        self._btn(ab2, "🔄  RESET", self.accent_purple, 'white',
                  self._bms_cfg_reset_defaults, w=14, h=2).pack(side=tk.LEFT, padx=4)
        self._btn(ab2, "💾  SAVE JSON", '#334155', 'white',
                  self._bms_cfg_save_json, w=14, h=2).pack(side=tk.LEFT, padx=4)

        prog_frame = tk.Frame(c, bg=self.bg_card); prog_frame.pack(fill=tk.X, pady=(0,6))
        pf2 = tk.Frame(prog_frame, bg=self.bg_card); pf2.pack(fill=tk.X, padx=16, pady=8)
        tk.Label(pf2, text="Progress:", font=self.font_small, bg=self.bg_card,
                 fg=self.text_secondary).pack(side=tk.LEFT)
        self.bms_progress_bar = ttk.Progressbar(pf2, length=450, mode='determinate',
                                                  maximum=100, value=0)
        self.bms_progress_bar.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        self.bms_progress_lbl = tk.Label(pf2, text="—", font=self.font_small,
                                          bg=self.bg_card, fg=self.text_secondary)
        self.bms_progress_lbl.pack(side=tk.LEFT)

        tbl_frame = tk.Frame(c, bg=self.bg_card2)
        tbl_frame.pack(fill=tk.BOTH, expand=True, pady=(0,6))
        hdr_f = tk.Frame(tbl_frame, bg=self.accent_blue, height=26)
        hdr_f.pack(fill=tk.X, padx=2, pady=(2,0)); hdr_f.pack_propagate(False)
        for txt, w in [("☑","3"),("Signal Name","22"),("Address","10"),
                       ("Type","6"),("Value","12"),("Min","8"),("Max","8"),("Status","12")]:
            tk.Label(hdr_f, text=txt, width=int(w), font=('Consolas',9,'bold'),
                     bg=self.accent_blue, fg='black', anchor=tk.W).pack(side=tk.LEFT, padx=2)
        body_frame = tk.Frame(tbl_frame, bg=self.bg_card2)
        body_frame.pack(fill=tk.BOTH, expand=True, padx=2)
        canvas = tk.Canvas(body_frame, bg='#080e22', highlightthickness=0)
        vsb    = tk.Scrollbar(body_frame, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y); canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._bms_tbl_inner = tk.Frame(canvas, bg='#080e22')
        self._bms_tbl_inner_id = canvas.create_window((0,0), window=self._bms_tbl_inner, anchor='nw')
        self._bms_tbl_inner.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(-1*(e.delta//120), "units"))
        self._bms_canvas = canvas

        tk.Label(c, text="📋  CONFIG LOG", font=self.font_header,
                 bg=self.bg_dark, fg=self.accent_blue).pack(anchor=tk.W)
        self.bms_cfg_log = scrolledtext.ScrolledText(c, height=6, bg='#080e22',
                                                     fg='#ffd966', font=self.font_mono)
        self.bms_cfg_log.pack(fill=tk.X, pady=(4,0))
        for tag, fg in [("ok","#00ff95"),("err","#ff4d4d"),
                        ("inf","#ffd966"),("retry","#ffaa00")]:
            self.bms_cfg_log.tag_config(tag, foreground=fg)
        self._bms_row_widgets = []; self._bms_check_vars = []

    # ------------------------------------------------------------------
    # RAW DATA TAB
    # ------------------------------------------------------------------

    def _build_raw_tab(self, parent):
        card = self._card(parent, "📨  RAW HEX DATA", self.accent_blue)
        card.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        hf = tk.Frame(card, bg=self.bg_card); hf.pack(fill=tk.X, padx=16, pady=(0,8))
        self._btn(hf, "CLEAR", self.accent_red, 'white',
                  self._clear_raw, w=8, h=1).pack(side=tk.RIGHT, padx=4)
        self.pause_raw_btn = self._btn(hf, "PAUSE", self.accent_orange, 'white',
                                        self._toggle_raw_pause, w=8, h=1)
        self.pause_raw_btn.pack(side=tk.RIGHT, padx=4)
        self.raw_frame_count_lbl = tk.Label(hf, text="Frames: 0", font=self.font_small,
                                            bg=self.bg_card, fg=self.text_secondary)
        self.raw_frame_count_lbl.pack(side=tk.RIGHT, padx=12)
        self.raw_hex_text = scrolledtext.ScrolledText(
            card, bg='#080e22', fg='#00ff9d', font=('Courier New',9,'bold'),
            wrap=tk.NONE)
        self.raw_hex_text.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0,16))
        for tag, fg in [("bat","#00d4ff"),("sd","#00ff95"),
                        ("cfg","#ffd966"),("ota","#ffaa00"),("other","#8899aa")]:
            self.raw_hex_text.tag_config(tag, foreground=fg)
        self._raw_paused = False; self._raw_frame_count = 0

    # ==================================================================
    # HELPER WIDGETS
    # ==================================================================

    def _card(self, parent, title: str, accent: str) -> tk.Frame:
        outer = tk.Frame(parent, bg=self.bg_card, relief='flat')
        hdr   = tk.Frame(outer, bg=self.bg_card)
        hdr.pack(fill=tk.X)
        accent_bar = tk.Frame(hdr, bg=accent, width=4)
        accent_bar.pack(side=tk.LEFT, fill=tk.Y)
        tk.Label(hdr, text=f"  {title}", font=self.font_header,
                 bg=self.bg_card, fg=accent).pack(side=tk.LEFT, pady=10)
        return outer

    def _btn(self, parent, text, bg, fg, cmd, w=14, h=2, state=tk.NORMAL):
        return tk.Button(parent, text=text, font=self.font_small, bg=bg, fg=fg,
                         width=w, height=h, command=cmd, state=state,
                         relief='flat', activebackground=bg, cursor='hand2')

    # ==================================================================
    # OTA UI CALLBACKS
    # ==================================================================

    def _ota_browse(self):
        path = filedialog.askopenfilename(
            title="Select firmware .bin file",
            filetypes=[("Binary files","*.bin"),("All files","*.*")])
        if not path: return
        self.ota_file_var.set(path)
        size = os.path.getsize(path)
        remainder = size % CHUNK_SIZE
        padded_size = size + (CHUNK_SIZE - remainder) if remainder else size
        with open(path, "rb") as f: data = f.read()
        padded = data + bytes(CHUNK_SIZE - remainder) if remainder else data
        crc    = crc32_firmware(padded)
        n_chunks = padded_size // CHUNK_SIZE
        self.ota_file_info["Path:"].config(text=os.path.basename(path))
        self.ota_file_info["Size:"].config(text=f"{size} B  ({size/1024:.2f} KB)")
        self.ota_file_info["Padded Size:"].config(
            text=f"{padded_size} B  ({n_chunks} × {CHUNK_SIZE}B chunks)")
        self.ota_file_info["CRC32 (reflected):"].config(
            text=f"0x{crc:08X}  (poly 0xEDB88320)")
        if self.ble.data.connected:
            self.ota_start_btn.config(state=tk.NORMAL)
        self._ota_log(f"📁 {os.path.basename(path)}  {size}B  CRC=0x{crc:08X}", "info")

    def _start_ota(self):
        fp = self.ota_file_var.get().strip()
        if not fp or not os.path.exists(fp):
            messagebox.showerror("Error", "Select a valid .bin firmware file first"); return
        if not self.ble.data.connected:
            messagebox.showerror("Error", "Not connected to device"); return
        size     = os.path.getsize(fp)
        n_chunks = (size + CHUNK_SIZE - 1) // CHUNK_SIZE
        if not messagebox.askyesno("Confirm OTA Update",
            f"Start full OTA firmware update?\n\n"
            f"File : {os.path.basename(fp)}\n"
            f"Size : {size} bytes\n\n"
            f"Sequence:\n"
            f"  1. TRIGGER  (0x36)\n"
            f"  2. FLASH    (0x33) — erase flash\n"
            f"  3. SIZE     (0x01) — send size\n"
            f"  4. CHUNKS   (0x02) — {n_chunks} chunks\n"
            f"  5. CRC      (0x03) — verify & reboot\n\n"
            f"This will ERASE and REPLACE firmware."): return

        # Reset all step indicators
        for key in self.ota_step_labels:
            self._ota_set_step(key, 'idle')
        self._ota_ack_count = 0
        self.ota_progress_bar['value'] = 0
        self.ota_progress_bar['maximum'] = 100
        self.ota_progress_lbl.config(text="0 / 0")
        self.ota_pct_lbl.config(text="0%", fg=self.accent_cyan)
        self.ota_chunk_lbl.config(text="—")
        self.ota_status_lbl.config(text="OTA in progress…", fg=self.accent_orange)
        self.ota_start_btn.config(state=tk.DISABLED)
        self.ble.request_ota(fp)
        self._add_log(f"🚀 OTA started: {os.path.basename(fp)}")

    def _ota_log(self, msg: str, tag: str = "info"):
        ts = datetime.now().strftime("%H:%M:%S")
        self.ota_log_text.insert(tk.END, f"[{ts}] {msg}\n", tag)
        self.ota_log_text.see(tk.END)

    def _ota_set_step(self, key: str, state: str):
        """state: 'idle' | 'active' | 'done' | 'fail'"""
        cfg = {
            'idle':   ("○", self.text_secondary),
            'active': ("▶", self.accent_yellow),
            'done':   ("✔", self.accent_green),
            'fail':   ("✗", self.accent_red),
        }
        icon, color = cfg.get(state, ("○", self.text_secondary))
        if key in self.ota_step_labels:
            self.ota_step_labels[key].config(text=icon, fg=color)
        self._ota_step_state[key] = state

    # ==================================================================
    # BMS CONFIG CALLBACKS
    # ==================================================================

    def _bms_cfg_browse(self):
        path = filedialog.askopenfilename(
            title="Select BMS Config Excel",
            filetypes=[("Excel files","*.xlsx *.xls"),("All files","*.*")])
        if path: self._bms_cfg_excel_path.set(path)

    def _bms_cfg_load(self):
        path = self._bms_cfg_excel_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("Error", "Please select a valid Excel file first."); return
        if not _PANDAS_OK:
            messagebox.showerror("Error", "pandas is required: pip install pandas openpyxl"); return
        try:
            import pandas as pd
            df = pd.read_excel(path, sheet_name='BMS_CONFIG_V3', header=1)
            df = df[pd.notna(df['Signal_Name'])].reset_index(drop=True)
            self._bms_params = []
            for _, row in df.iterrows():
                addr_str = str(row['Address']).strip()
                addr_int = int(addr_str, 16) if addr_str.startswith('0x') else int(addr_str)
                calc_val = row['CalculatedValue']
                if isinstance(calc_val, str) and calc_val.startswith('='): calc_val = 0.0
                else:
                    try: calc_val = float(calc_val)
                    except: calc_val = 0.0
                self._bms_params.append({
                    'signal_name': str(row['Signal_Name']).strip(),
                    'address'    : addr_int,
                    'length'     : int(row['Length']),
                    'data_type'  : str(row['Data_Type']).strip(),
                    'default_val': calc_val,
                    'value'      : calc_val,
                    'min_val'    : float(row['MinValue']) if pd.notna(row['MinValue']) else None,
                    'max_val'    : float(row['MaxValue']) if pd.notna(row['MaxValue']) else None,
                })
            self._bms_populate_table()
            self.bms_send_all_btn.config(state=tk.NORMAL)
            self.bms_send_sel_btn.config(state=tk.NORMAL)
            self._bms_cfg_log_add(
                f"✅ Loaded {len(self._bms_params)} params from {os.path.basename(path)}", "ok")
        except Exception as e:
            messagebox.showerror("Load Error", str(e))
            self._bms_cfg_log_add(f"❌ Load failed: {e}", "err")

    def _bms_populate_table(self):
        for w in self._bms_tbl_inner.winfo_children(): w.destroy()
        self._bms_row_widgets = []; self._bms_check_vars = []
        for idx, p in enumerate(self._bms_params):
            row_bg = '#0d1229' if idx % 2 == 0 else '#0a1020'
            rf = tk.Frame(self._bms_tbl_inner, bg=row_bg, height=26)
            rf.pack(fill=tk.X, pady=1); rf.pack_propagate(False)
            chk_var = tk.BooleanVar(value=True); self._bms_check_vars.append(chk_var)
            tk.Checkbutton(rf, variable=chk_var, bg=row_bg,
                           activebackground=row_bg).pack(side=tk.LEFT, padx=2)
            tk.Label(rf, text=p['signal_name'], width=22, anchor=tk.W,
                     font=self.font_mono, bg=row_bg, fg=self.accent_blue).pack(side=tk.LEFT)
            tk.Label(rf, text=f"0x{p['address']:08X}", width=10, anchor=tk.W,
                     font=self.font_mono, bg=row_bg, fg=self.text_secondary).pack(side=tk.LEFT)
            dt_col = {'U32':self.accent_green,'I32':self.accent_orange,
                      'I16':self.accent_yellow,'I8':self.accent_yellow}.get(
                          p['data_type'], self.text_primary)
            tk.Label(rf, text=p['data_type'], width=6, anchor=tk.W,
                     font=self.font_mono, bg=row_bg, fg=dt_col).pack(side=tk.LEFT)
            val_var = tk.StringVar(value=str(p['value']))
            entry = tk.Entry(rf, textvariable=val_var, width=12, bg='#0a1628',
                             fg=self.text_primary, font=self.font_mono, relief='flat')
            entry.pack(side=tk.LEFT, padx=2)
            for txt in [str(p['min_val']) if p['min_val'] is not None else "—",
                        str(p['max_val']) if p['max_val'] is not None else "—"]:
                tk.Label(rf, text=txt, width=8, anchor=tk.W,
                         font=self.font_mono, bg=row_bg, fg=self.text_secondary).pack(side=tk.LEFT)
            status_lbl = tk.Label(rf, text="—", width=12, anchor=tk.W,
                                  font=self.font_mono, bg=row_bg, fg=self.text_secondary)
            status_lbl.pack(side=tk.LEFT)
            self._bms_row_widgets.append({'val_var':val_var,'entry':entry,
                                          'status_lbl':status_lbl,'row_bg':row_bg})
        self._bms_tbl_inner.update_idletasks()
        self._bms_canvas.configure(scrollregion=self._bms_canvas.bbox("all"))

    def _bms_cfg_send_all(self):      self._bms_cfg_do_send(send_all=True)
    def _bms_cfg_send_selected(self): self._bms_cfg_do_send(send_all=False)

    def _bms_cfg_do_send(self, send_all: bool):
        if not self._bms_params:
            messagebox.showwarning("No Data", "Load an Excel file first."); return
        if not self.ble.data.connected:
            messagebox.showerror("Not Connected", "Connect to BMS device first."); return
        to_send = []
        for idx, (p, rw) in enumerate(zip(self._bms_params, self._bms_row_widgets)):
            if not send_all and not self._bms_check_vars[idx].get(): continue
            raw_val = rw['val_var'].get().strip()
            try: val = float(raw_val) if '.' in raw_val else int(raw_val)
            except ValueError:
                rw['status_lbl'].config(text="INVALID", fg=self.accent_red)
                self._bms_cfg_log_add(f"❌ {p['signal_name']}: invalid '{raw_val}'", "err"); continue
            if p['min_val'] is not None and float(val) < p['min_val']:
                rw['status_lbl'].config(text="BELOW MIN", fg=self.accent_orange)
                self._bms_cfg_log_add(
                    f"⚠ {p['signal_name']}: {val} < min {p['min_val']} — skipped", "err"); continue
            if p['max_val'] is not None and float(val) > p['max_val']:
                rw['status_lbl'].config(text="ABOVE MAX", fg=self.accent_orange)
                self._bms_cfg_log_add(
                    f"⚠ {p['signal_name']}: {val} > max {p['max_val']} — skipped", "err"); continue
            rw['status_lbl'].config(text="PEND", fg=self.accent_yellow)
            to_send.append({'signal_name':p['signal_name'],'address':p['address'],
                            'data_type':p['data_type'],'value':val,'_idx':idx})
        if not to_send:
            messagebox.showwarning("Nothing to Send", "No valid parameters to send."); return
        total = len(to_send)
        self.bms_progress_bar['maximum'] = total
        self.bms_progress_bar['value']   = 0
        self.bms_progress_lbl.config(text=f"0 / {total}")
        self._bms_pending_count  = total
        self._bms_ack_received   = 0
        self._bms_ack_ok         = 0
        self._bms_ack_fail       = 0
        self._bms_cfg_log_add(f"📡 Sending {total} parameters (ACK-based)", "inf")
        self.ble.send_bms_config_all(to_send)

    def _bms_cfg_reset_defaults(self):
        for p, rw in zip(self._bms_params, self._bms_row_widgets):
            rw['val_var'].set(str(p['default_val']))
            rw['status_lbl'].config(text="—", fg=self.text_secondary)
        self._bms_cfg_log_add("🔄 Values reset to Excel defaults.", "inf")

    def _bms_cfg_save_json(self):
        import json
        if not self._bms_params:
            messagebox.showwarning("No Data", "Load parameters first."); return
        path = filedialog.asksaveasfilename(title="Save BMS Config",
                defaultextension=".json", filetypes=[("JSON files","*.json")])
        if not path: return
        data = [{'signal':p['signal_name'],'address':f"0x{p['address']:08X}",
                 'data_type':p['data_type'],'value':rw['val_var'].get()}
                for p, rw in zip(self._bms_params, self._bms_row_widgets)]
        with open(path, 'w') as f: import json; json.dump(data, f, indent=2)
        self._bms_cfg_log_add(f"💾 Saved to {os.path.basename(path)}", "ok")

    def _bms_cfg_log_add(self, msg: str, tag: str = "inf"):
        ts = datetime.now().strftime("%H:%M:%S")
        self.bms_cfg_log.insert(tk.END, f"[{ts}] {msg}\n", tag)
        self.bms_cfg_log.see(tk.END)

    # ==================================================================
    # UNIFIED BLE EVENT DISPATCHER
    # ==================================================================

    def _on_ble_event(self, event):
        if isinstance(event, BatteryData):
            self.root.after(0, lambda d=event: self._update_battery_display(d))
            return
        if not isinstance(event, tuple) or not event:
            return

        kind = event[0]

        def dispatch():
            if kind == "scan_started":
                self.scan_status.config(text="Scanning…")
            elif kind == "scan_results":
                self._tk_scan_results(event[1])
            elif kind == "scan_error":
                self.scan_status.config(text=f"Scan error: {event[1]}")
            elif kind == "status":
                self.scan_status.config(text=event[1])
            elif kind == "connected":
                self._tk_connected()
            elif kind == "streaming_status":
                self._tk_streaming_status(event[1])
            elif kind == "command_sent":
                self._add_log(f"📤 {event[1]}")
            elif kind == "command_response":
                self._add_log(f"📥 {event[1]}")
            elif kind == "raw_data":
                self._tk_raw(event[1])
            elif kind == "sd_card_info":
                self._tk_sd_card_info(event[1])
            elif kind == "sd_log_started":
                self._add_log(f"▶ {event[1]}")
            elif kind == "sd_read_started":
                self._tk_sd_read_started(event[1])
            elif kind == "sd_subframe":
                self._tk_sd_subframe(event[1])
            elif kind == "sd_decode_error":
                self._tk_sd_decode_error(event[1])
            elif kind == "sd_read_complete":
                self._tk_sd_complete(event[1])
            elif kind == "sd_erase_result":
                self._tk_sd_erase_result(event[1])
            elif kind == "sd_error":
                self._add_log(f"❌ SD: {event[1]}")
            elif kind == "sd_timeout":
                self._tk_sd_timeout(event[1])
            elif kind == "bms_cfg_started":
                self._tk_bms_cfg_started(event[1])
            elif kind == "bms_cfg_progress":
                self._tk_bms_cfg_progress(event[1])
            elif kind == "bms_cfg_retry":
                self._tk_bms_cfg_retry(event[1])
            elif kind == "bms_cfg_timeout_retry":
                self._tk_bms_cfg_timeout_retry(event[1])
            elif kind == "bms_cfg_done":
                self._tk_bms_cfg_done(event[1])
            elif kind == "bms_cfg_ack":
                self._tk_bms_cfg_ack(event[1])
            elif kind == "bms_cfg_timeout":
                self._tk_bms_cfg_timeout(event[1])
            elif kind == "bms_cfg_error":
                self._bms_cfg_log_add(f"❌ {event[1]}", "err")
            # OTA events
            elif kind == "ota_started":
                file_size    = event[1] if len(event) > 1 else 0
                total_chunks = event[2] if len(event) > 2 else 0
                self._tk_ota_started(file_size, total_chunks)
            elif kind == "ota_progress":
                chunk = event[1] if len(event) > 1 else 0
                total = event[2] if len(event) > 2 else 0
                pct   = event[3] if len(event) > 3 else 0
                self._tk_ota_progress(chunk, total, pct)
            elif kind == "ota_step":
                key   = event[1] if len(event) > 1 else ""
                state = event[2] if len(event) > 2 else "idle"
                self._ota_set_step(key, state)
            elif kind == "ota_complete":
                self._tk_ota_complete(event[1])
            elif kind == "ota_log":
                self._tk_ota_log(event[1])
            elif kind == "ota_ack_raw":
                self._tk_ota_ack_raw(event[1])
            elif kind == "ota_error":
                self._ota_log(f"❌ {event[1]}", "err")

        self.root.after(0, dispatch)

    # OTA GUI handlers
    def _tk_ota_started(self, file_size: int, total_chunks: int):
        self._ota_total_chunks = total_chunks
        self.ota_progress_bar['maximum'] = total_chunks if total_chunks > 0 else 100
        self.ota_progress_bar['value']   = 0
        self.ota_progress_lbl.config(text=f"0 / {total_chunks}")
        self.ota_pct_lbl.config(text="0%", fg=self.accent_cyan)
        self.ota_chunk_lbl.config(text=f"0 / {total_chunks} chunks")
        self.ota_status_lbl.config(text="OTA in progress…", fg=self.accent_orange)
        self.ota_start_btn.config(state=tk.DISABLED)
        self._ota_log(f"🚀 OTA: {file_size}B  {total_chunks} chunks", "info")

    def _tk_ota_progress(self, chunk: int, total: int, pct: int):
        self.ota_progress_bar['maximum'] = total if total > 0 else 100
        self.ota_progress_bar['value']   = chunk
        self.ota_progress_lbl.config(text=f"{chunk} / {total}")
        self.ota_pct_lbl.config(text=f"{pct}%",
                                  fg=(self.accent_green if pct >= 100 else self.accent_cyan))
        self.ota_chunk_lbl.config(text=f"{chunk} / {total} chunks written")

    def _tk_ota_complete(self, success: bool):
        self.ota_start_btn.config(
            state=tk.NORMAL if (self.ota_file_var.get() and self.ble.data.connected)
                  else tk.DISABLED)
        if success:
            self.ota_progress_bar['value'] = self.ota_progress_bar['maximum']
            self.ota_pct_lbl.config(text="100%", fg=self.accent_green)
            self.ota_status_lbl.config(text="✔ OTA COMPLETE — device rebooting",
                                        fg=self.accent_green)
            for key in self.ota_step_labels:
                if self._ota_step_state.get(key) != 'fail':
                    self._ota_set_step(key, 'done')
            self._ota_log("✅ OTA SUCCESS — device is rebooting with new firmware", "ok")
            self._add_log("✅ OTA COMPLETE")
            messagebox.showinfo("OTA Complete",
                                "Firmware updated successfully!\nDevice is rebooting.")
        else:
            self.ota_status_lbl.config(text="✗ OTA FAILED", fg=self.accent_red)
            self.ota_pct_lbl.config(text="FAIL", fg=self.accent_red)
            self._ota_log("❌ OTA FAILED — check log above for details", "err")
            self._add_log("❌ OTA FAILED")
            messagebox.showerror("OTA Failed", "Firmware update failed.\nCheck the OTA log.")

    def _tk_ota_log(self, msg: str):
        tag = "info"
        if "ACK OK" in msg: tag = "ok"
        if "FAIL" in msg or "failed" in msg.lower(): tag = "err"
        if "[TX " in msg: tag = "warning"
        self._ota_log(msg, tag)

    def _tk_ota_ack_raw(self, ok: bool):
        self._ota_ack_count += 1
        ts  = datetime.now().strftime("%H:%M:%S")
        sym = "✓" if ok else "✗"
        tag = "ok" if ok else "err"
        self.ota_ack_hist.insert(tk.END, f"[{ts}] #{self._ota_ack_count} {sym}\n", tag)
        self.ota_ack_hist.see(tk.END)
        lines = int(self.ota_ack_hist.index('end-1c').split('.')[0])
        if lines > 20:
            self.ota_ack_hist.delete('1.0', f'{lines-20}.0')

    # BMS config GUI callbacks
    def _tk_bms_cfg_started(self, total: int):
        self.bms_progress_bar['maximum'] = total if total > 0 else 1
        self.bms_progress_bar['value']   = 0
        self.bms_progress_lbl.config(text=f"0 / {total} — waiting for ACK…")
        self._bms_ack_received = 0
        self._bms_ack_ok       = 0
        self._bms_ack_fail     = 0
        self._bms_pending_count = total

    def _tk_bms_cfg_progress(self, info):
        sent, total, signal_name, ok, retry = info
        self.bms_progress_bar['maximum'] = total if total > 0 else 1
        self.bms_progress_bar['value']   = sent
        self.bms_progress_lbl.config(text=f"{sent} / {total} — {signal_name}")
        for idx, p in enumerate(self._bms_params):
            if p['signal_name'] == signal_name and idx < len(self._bms_row_widgets):
                rw = self._bms_row_widgets[idx]
                rw['status_lbl'].config(
                    text="✔ ACK OK" if ok else "✗ FAIL",
                    fg=self.accent_green if ok else self.accent_red)
                break

    def _tk_bms_cfg_retry(self, info):
        param_name, retry, max_retries = info
        self._bms_cfg_log_add(f"⟳ Retry {retry}/{max_retries}: {param_name}", "retry")
        for idx, p in enumerate(self._bms_params):
            if p['signal_name'] == param_name and idx < len(self._bms_row_widgets):
                self._bms_row_widgets[idx]['status_lbl'].config(
                    text=f"RETRY {retry}", fg=self.accent_orange); break

    def _tk_bms_cfg_timeout_retry(self, info):
        param_name, retry, max_retries = info
        self._bms_cfg_log_add(f"⏱ Timeout {param_name} retry {retry}/{max_retries}", "retry")
        for idx, p in enumerate(self._bms_params):
            if p['signal_name'] == param_name and idx < len(self._bms_row_widgets):
                self._bms_row_widgets[idx]['status_lbl'].config(
                    text=f"TIMEOUT {retry}", fg=self.accent_orange); break

    def _tk_bms_cfg_done(self, info):
        success, total, failed = info
        self.bms_progress_bar['value'] = total
        self.bms_progress_lbl.config(
            text=f"✔ {success}/{total} ACKs" +
                 (f"  ({len(failed)} failed)" if failed else ""))
        self._bms_cfg_log_add(
            f"📡 Done: {success}/{total} OK. "
            f"{'COMPLETE' if success==total else 'PARTIAL' if success>0 else 'FAILED'}.",
            "ok" if success==total else "inf")
        if failed:
            self._bms_cfg_log_add(f"   Failed: {', '.join(failed[:10])}", "err")
        self._add_log(f"⚙️ BMS Config: {success}/{total} ACKs")

    def _tk_bms_cfg_ack(self, ok: bool):
        self._bms_ack_received += 1
        if ok: self._bms_ack_ok   += 1
        else:  self._bms_ack_fail += 1
        total = self._bms_pending_count
        self.bms_progress_lbl.config(
            text=f"ACK {self._bms_ack_received}/{total}  "
                 f"✔{self._bms_ack_ok}  ✗{self._bms_ack_fail}")
        if not ok:
            self._bms_cfg_log_add(
                f"✗ ACK #{self._bms_ack_received}: FAILED", "err")

    def _tk_bms_cfg_timeout(self, timeout_params):
        if timeout_params:
            self._bms_cfg_log_add(
                f"⏱ Timeout: {', '.join(timeout_params[:5])}", "err")
            for param_name in timeout_params:
                for idx, p in enumerate(self._bms_params):
                    if p['signal_name'] == param_name and idx < len(self._bms_row_widgets):
                        self._bms_row_widgets[idx]['status_lbl'].config(
                            text="TIMEOUT", fg=self.accent_red); break

    # General GUI callbacks
    def _tk_scan_results(self, devices):
        for item in self.device_tree.get_children(): self.device_tree.delete(item)
        for i, d in enumerate(devices, 1):
            self.device_tree.insert('','end', text=str(i),
                values=(d['name'], d['address'], f"{d['rssi']} dBm"))
        self.scan_status.config(text=f"Found {len(devices)} device(s)")

    def _tk_connected(self):
        self.connect_btn.config(state=tk.DISABLED)
        self.disconnect_btn.config(state=tk.NORMAL)
        for btn in (self.sd_info_btn, self.sd_read_btn, self.sd_erase_btn):
            btn.config(state=tk.NORMAL)
        if self.ota_file_var.get():
            self.ota_start_btn.config(state=tk.NORMAL)
        self.scan_status.config(text="Connected")
        self._add_log("✅ Connected — battery streaming started")

    def _tk_streaming_status(self, status: str):
        if status == "STREAMING":
            self.stream_status.config(text="▶ STREAMING: ON", fg=self.accent_green)
        elif status == "PAUSED":
            self.stream_status.config(text="⏸ STREAMING: PAUSED", fg=self.accent_orange)
        else:
            self.stream_status.config(text="▶ STREAMING: OFF", fg=self.text_secondary)

    def _tk_raw(self, data_bytes: bytes):
        if self._raw_paused: return
        self._raw_frame_count += 1
        ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        is_framed = (len(data_bytes) >= 4 and
                     data_bytes[0] == SOF_BYTE_0 and data_bytes[1] == SOF_BYTE_1)
        if is_framed:
            cmd     = data_bytes[3]
            cmd_str = {CMD_BAT_START:"BAT_START", CMD_BAT_STOP:"BAT_STOP",
                       CMD_SD_DATA:"SD_DATA", CMD_BMS_CONFIG:"BMS_CFG",
                       CMD_OTA_UPDATE:"OTA"}.get(cmd, f"0x{cmd:02X}")
            tag     = ("bat" if cmd in (CMD_BAT_START, CMD_BAT_STOP)
                       else "cfg" if cmd == CMD_BMS_CONFIG
                       else "ota" if cmd == CMD_OTA_UPDATE else "sd")
        else:
            cmd_str = "???"; tag = "other"
        hex_s = ' '.join(f'{b:02X}' for b in data_bytes)
        self.raw_hex_text.insert(tk.END,
            f"[{ts}] {cmd_str:<10} Len:{len(data_bytes):4d} | {hex_s}\n", tag)
        self.raw_hex_text.see(tk.END)
        self.raw_frame_count_lbl.config(text=f"Frames: {self._raw_frame_count}")
        lines = int(self.raw_hex_text.index('end-1c').split('.')[0])
        if lines > 200:
            self.raw_hex_text.delete('1.0', f'{lines-200}.0')

    def _tk_sd_card_info(self, info):
        ctypes = {0x01:"SDSC",0x02:"SDHC"}
        self.sd_info_text.delete(1.0, tk.END)
        ct = ctypes.get(info['card_type'], f"0x{info['card_type']:02X}")
        self.sd_info_text.insert(tk.END,
            f"Card Type         : {ct}\n"
            f"Capacity          : {info['capacity_gb']} GB\n"
            f"Total Sectors     : {info['total_sectors']:,}\n"
            f"Remaining Sectors : {info['remaining_sectors']:,}\n"
            f"Current Sector    : {info['current_sector']:,} (0x{info['current_sector']:08X})\n"
            f"Manufacture Date  : {info['manufacture_month']:02d}/{info['manufacture_year']}\n"
            f"Manufacturer ID   : 0x{info['manufacturer_id']:02X}\n")
        self.sd_status_lbl.config(text="Card info received ✔")

    def _tk_sd_read_started(self, info):
        start_sec, end_sec, expected, file_path = info
        self.sd_counter_lbl.config(text=f"Records: 0 / {expected}")
        self.sd_progress_bar['maximum'] = expected if expected > 0 else 1
        self.sd_progress_bar['value']   = 0
        self.sd_progress_lbl.config(text=f"0 / {expected}")
        self.sd_log_indicator.config(text="⬤ LOGGING", fg=self.accent_green)
        self.sd_status_lbl.config(text=f"Reading sectors {start_sec}–{end_sec}")
        self.sd_data_text.delete(1.0, tk.END)
        self.sd_data_text.insert(tk.END,
            f"📖 READ: {start_sec}→{end_sec}  ({expected} frames)  {file_path}\n"
            f"{'─'*60}\n\n", "header")

    def _tk_sd_subframe(self, info):
        received, expected, records, signals, file_path, ok = info
        self.sd_counter_lbl.config(text=f"Records: {records} / {expected}")
        self.sd_progress_bar['value'] = received
        self.sd_progress_lbl.config(text=f"{received} / {expected}")
        self.sd_data_text.insert(tk.END,
            f"\n─── Record {records} ({received}/{expected}) ───\n", "header")
        self.sd_data_text.insert(tk.END,
            f"SOC:{signals.get('SOC_percent',0)}%  "
            f"V:{signals.get('PackVoltage_V',0):.2f}  "
            f"A:{signals.get('PackCurrent_A',0):.2f}  "
            f"State:{signals.get('BMSOperationState',0)}\n", "data")
        self.sd_data_text.see(tk.END)

    def _tk_sd_decode_error(self, frame_num):
        self.sd_data_text.insert(tk.END, f"❌ Decode failed frame #{frame_num}\n", "error")
        self.sd_data_text.see(tk.END)

    def _tk_sd_complete(self, info):
        records, msg, file_path = info
        self.sd_log_indicator.config(text="⬤ DONE", fg=self.accent_yellow)
        self.sd_status_lbl.config(text=f"✅ Complete — {records} records")
        self.sd_data_text.insert(tk.END,
            f"\n{'═'*60}\n✅ {records} records  {file_path}\n{'═'*60}\n", "done")
        self.sd_data_text.see(tk.END)
        self._add_log(f"✅ SD read complete: {msg}")

    def _tk_sd_erase_result(self, info):
        success, start_sec, end_sec = info
        n = end_sec - start_sec + 1
        msg = (f"✅ Erase SUCCESS: {start_sec:,}–{end_sec:,} ({n:,} sectors)"
               if success else f"❌ Erase FAILED: {start_sec:,}–{end_sec:,}")
        self.sd_data_text.insert(tk.END, f"\n{msg}\n", "done" if success else "error")
        self.sd_status_lbl.config(
            text=f"{'✅ Erased' if success else '❌ FAILED'} {n:,} sectors")
        self._add_log(msg)

    def _tk_sd_timeout(self, message: str):
        self.sd_log_indicator.config(text="⬤ TIMEOUT", fg=self.accent_red)
        self.sd_status_lbl.config(text="⏱ Timeout", fg=self.accent_red)
        self.sd_data_text.insert(tk.END, f"\n⏱ SD TIMEOUT — {message}\n", "timeout")
        self.sd_data_text.see(tk.END)
        self._add_log(f"⏱ {message}")
        self.root.after(3000, lambda: self.sd_status_lbl.config(fg=self.accent_yellow))

    def _update_battery_display(self, d: BatteryData):
        try:
            self.conn_status.config(
                text="● CONNECTED" if d.connected else "● DISCONNECTED",
                fg=self.accent_green if d.connected else self.accent_red)
            self.soc_label.config(text=str(d.SOC))
            self._draw_soc_bar(d.SOC)
            self.soh_label.config(text=f"{d.SOH}%")
            self.sop_label.config(text=f"{d.SOP}%")
            self.state_label.config(text=d.get_operation_state_string())
            self.fet_label.config(text=f"0x{d.FETStatus:02X}")
            self.chg_flag.config( fg=self.accent_green if d.BatteryChgFlag else self.text_secondary)
            self.key_flag.config( fg=self.accent_green if d.KeyOn          else self.text_secondary)
            self.comm_flag.config(fg=self.accent_green if d.BCUCommStatus  else self.text_secondary)
            self.chg_cap_label.config(  text=f"{d.ChargeCapacity:.2f} Ah")
            self.dchg_cap_label.config( text=f"{d.DischargeCapacity:.2f} Ah")
            self.socdiff_label.config(  text=f"{d.SOCDiff}%")
            self.pack_volt_label.config(text=f"{d.PackVoltage:.2f} V")
            self.pack_curr_label.config(text=f"{d.PackCurrent:.2f} A")
            self.fet_volt_label.config( text=f"{d.BattVoltageAfterFET:.3f} V")
            self.load_volt_label.config(text=f"{d.LoadVoltage:.2f} V")
            self.stack_volt_label.config(text=f"{d.StackVoltage:.3f} V")
            for i, lbl in enumerate(self.batt_temp_labels):
                t = d.BattTemperatures[i]
                lbl.config(text=f"{t:.1f}°C", fg=self._temp_color(t))
            for i, lbl in enumerate(self.pcb_temp_labels):
                t = d.PCBTemperatures[i]
                lbl.config(text=f"{t:.1f}°C", fg=self._temp_color(t))
            self.max_temp_badge.config(  text=f"{d.MaxBattTemp:.1f}°C")
            self.min_temp_badge.config(  text=f"{d.MinBattTemp:.1f}°C")
            self.delta_temp_badge.config(text=f"{d.CellTempDiff:.1f}°C")
            if d.MaxBattTemp >= self.temp_critical:
                self.temp_status.config(text="🔥 CRITICAL!", fg=self.accent_red)
                self.temp_warn_lbl.config(text=f"MAX:{d.MaxBattTemp:.1f}°C")
            elif d.MaxBattTemp >= self.temp_warning:
                self.temp_status.config(text="⚠ High Temp", fg=self.accent_orange)
                self.temp_warn_lbl.config(text=f"MAX:{d.MaxBattTemp:.1f}°C")
            else:
                self.temp_status.config(text="✓ Normal", fg=self.accent_green)
                self.temp_warn_lbl.config(text="")
            for i, lbl in enumerate(self.cell_labels):
                v = d.CellVoltages[i]
                lbl.config(text=f"{v:.3f}V",
                           fg=self.accent_red if (v < 3.0 or v > 4.2) else self.accent_green)
            self.max_cell_label.config( text=f"{d.MaxCellVoltage:.3f}V")
            self.min_cell_label.config( text=f"{d.MinCellVoltage:.3f}V")
            self.diff_cell_label.config(text=f"{d.CellVoltageDiff} mV")
            bt = f"0x{d.CellBalancingStatus:04X}"
            bt += " (Balancing)" if d.CellBalancingStatus else " (Idle)"
            self.balance_label.config(text=bt)
            for key, val in [
                ('min_cell_uv',d.MinCellUVFault),('max_cell_ov',d.MaxCellOVFault),
                ('over_chg_curr',d.OverChgCurrentFault),('over_dchg_curr',d.OverDchgCurrentFault),
                ('batt_ut',d.BattUTFault),('batt_ot',d.BattOTFault),
                ('batt_ov',d.BattOVFault),('batt_uv',d.BattUVFault)]:
                if key in self.fault_labels:
                    self.fault_labels[key].config(
                        text=d.get_fault_level_string(val), fg=d.get_fault_color(val))
            self.thermal_fault_label.config(
                fg=self.accent_red if d.ThermalRunawayFault else self.accent_green)
            for reason, key in [
                ("Min Cell UV",'MinCellUV'),("Max Cell Temp",'MaxCellTemp'),
                ("Max Cell Temp Rise",'MaxCellTempRise'),("Cell Temp Diff",'CellTempDiff'),
                ("Max Cell OV",'MaxCellOV'),("Cell Volt Sampling",'CellVolSamplingFault'),
                ("Temp Sampling",'TempSamplingFault')]:
                if reason in self.thermal_labels:
                    active = d.ThermalRunawayReasons[key]
                    self.thermal_labels[reason].config(
                        text="⚠" if active else "✓",
                        fg=self.accent_red if active else self.accent_green)
            for alarm, key in [
                ("Over PCB Temp",'OverPCBTemp'),("Pre-Charge Failure",'PreChargeFailure'),
                ("Hardware SC",'HardwareSC'),("Cell ΔV Alarm",'CellDVAlarm'),
                ("Cell ΔT Alarm",'CellTDAlarm'),("Low SOC Alarm",'LowSOCAlarm'),
                ("Low SOH Alarm",'LowSOHAlarm'),("SOC Jump Alarm",'SOCJumpAlarm')]:
                if alarm in self.alarm_labels:
                    active = d.Alarms[key]
                    self.alarm_labels[alarm].config(
                        text="⚠" if active else "✓",
                        fg=self.accent_red if active else self.accent_green)
            self.info_labels["Packet Count"].config(text=str(d.packet_count))
            self.info_labels["Last Update"].config(text=d.last_update)
            self.timestamp_label.config(text=f"Updated: {d.last_update}")
            self.notif_count_label.config(text=f"Pkt: {d.packet_count}")
        except Exception as e:
            print(f"GUI update error: {e}")

    def _temp_color(self, t):
        if t >= self.temp_critical: return self.accent_red
        if t >= self.temp_warning:  return self.accent_orange
        return self.accent_green

    # Button actions
    def _scan(self):
        self.scan_status.config(text="Scanning…"); self.ble.request_scan()

    def _connect(self):
        sel = self.device_tree.selection()
        if sel:
            item = self.device_tree.item(sel[0])
            self.ble.request_connect(item['values'][0], item['values'][1])
        else:
            name = self.device_name_var.get().strip()
            addr = self.device_addr_var.get().strip()
            if not name and not addr:
                messagebox.showerror("Error",
                    "Enter a device name/address or select one from the list"); return
            self.ble.request_connect(name or None, addr or None)

    def _disconnect(self):
        self.ble.request_disconnect()
        self.connect_btn.config(state=tk.NORMAL)
        self.disconnect_btn.config(state=tk.DISABLED)
        self.ota_start_btn.config(state=tk.DISABLED)
        for btn in (self.sd_info_btn, self.sd_read_btn, self.sd_erase_btn):
            btn.config(state=tk.DISABLED)
        self.stream_status.config(text="▶ STREAMING: OFF", fg=self.text_secondary)
        self._add_log("🛑 Disconnecting…")

    def _refresh_list(self):
        for item in self.device_tree.get_children(): self.device_tree.delete(item)
        for i, d in enumerate(self.ble.get_available_devices(), 1):
            self.device_tree.insert('','end', text=str(i),
                values=(d['name'], d['address'], f"{d['rssi']} dBm"))

    def _choose_bat_dir(self):
        d = filedialog.askdirectory(title="Select Battery Log Directory")
        if d:
            self.bat_log_dir = d
            self.log_status_labels['log_dir'].config(text=d)
            self._add_log(f"📁 Battery log dir: {d}")

    def _start_bat_log(self):
        ok, msg = self.bat_logger.start_logging(self.bat_log_dir)
        if ok:
            self.start_log_btn.config(state=tk.DISABLED)
            self.stop_log_btn.config(state=tk.NORMAL)
            self.log_status_labels['log_status'].config(text="ACTIVE", fg=self.accent_green)
            self.log_status_labels['log_file'].config(
                text=os.path.basename(self.bat_logger.file_path))
            self._update_log_timer()
            self._add_log(f"✅ {msg}")
        else:
            messagebox.showerror("Logging Error", msg)

    def _stop_bat_log(self):
        ok, msg = self.bat_logger.stop_logging()
        if ok:
            self.start_log_btn.config(state=tk.NORMAL)
            self.stop_log_btn.config(state=tk.DISABLED)
            self.log_status_labels['log_status'].config(text="Stopped", fg=self.accent_red)
            self._add_log(f"✅ {msg}")
        else:
            messagebox.showerror("Logging Error", msg)

    def _update_log_timer(self):
        if self.bat_logger.logging_active:
            self.log_status_labels['log_count'].config(text=str(self.bat_logger.log_count))
            if self.bat_logger.start_time:
                dur  = datetime.now() - self.bat_logger.start_time
                h, r = divmod(dur.seconds, 3600); m, s = divmod(r, 60)
                self.log_status_labels['log_duration'].config(text=f"{h:02d}:{m:02d}:{s:02d}")
            self.root.after(1000, self._update_log_timer)

    def _sd_select_dir(self):
        d = filedialog.askdirectory(title="Select SD Log Directory")
        if d:
            self.ble.set_sd_log_directory(d)
            self.sd_dir_lbl.config(text=d)
            self._add_log(f"📁 SD log dir: {d}")

    def _sd_info(self):
        self.ble.send_sd_info(); self.sd_status_lbl.config(text="Requesting card info…")

    def _sd_read_dialog(self):
        if not self.ble._sd_log_dir:
            messagebox.showwarning("No Directory", "Select SD log directory first"); return
        dlg = tk.Toplevel(self.root); dlg.title("Read Sectors"); dlg.geometry("400x200")
        dlg.configure(bg='#0d1229'); dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text="Read SD Card Sectors", fg=self.accent_green,
                 bg='#0d1229', font=('Consolas',13,'bold')).pack(pady=10)
        gf = tk.Frame(dlg, bg='#0d1229'); gf.pack(pady=8)
        entries = {}
        for i, (lbl, default) in enumerate([("Start Sector:","0"),("End Sector:","10")]):
            tk.Label(gf, text=lbl, fg=self.text_primary, bg='#0d1229',
                     font=self.font_normal).grid(row=i, column=0, padx=12, pady=5, sticky=tk.W)
            e = tk.Entry(gf, font=self.font_normal, width=14,
                         bg='#0a1628', fg=self.text_primary, relief='flat')
            e.insert(0, default); e.grid(row=i, column=1, padx=12, pady=5)
            entries[i] = e
        def do_read():
            try:
                start = int(entries[0].get()); end = int(entries[1].get())
                if end < start: messagebox.showerror("Error","End ≥ Start required"); return
                self.ble.send_sd_read(start, end); dlg.destroy()
            except ValueError: messagebox.showerror("Error","Invalid numbers")
        bf = tk.Frame(dlg, bg='#0d1229'); bf.pack(pady=12)
        self._btn(bf, "Send Request", self.accent_green, 'black',
                  do_read, w=14, h=2).pack(side=tk.LEFT, padx=5)
        self._btn(bf, "Cancel", self.accent_red, 'white',
                  dlg.destroy, w=10, h=2).pack(side=tk.LEFT, padx=5)

    def _sd_erase_dialog(self):
        dlg = tk.Toplevel(self.root); dlg.title("Erase Sectors"); dlg.geometry("380x200")
        dlg.configure(bg='#0d1229'); dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text="Erase SD Card Sectors", fg=self.accent_orange,
                 bg='#0d1229', font=('Consolas',13,'bold')).pack(pady=10)
        gf = tk.Frame(dlg, bg='#0d1229'); gf.pack(pady=8)
        entries = {}
        for i, (lbl, default) in enumerate([("Start Sector:","0"),("End Sector:","10")]):
            tk.Label(gf, text=lbl, fg=self.text_primary, bg='#0d1229',
                     font=self.font_normal).grid(row=i, column=0, padx=12, pady=5, sticky=tk.W)
            e = tk.Entry(gf, font=self.font_normal, width=14,
                         bg='#0a1628', fg=self.text_primary, relief='flat')
            e.insert(0, default); e.grid(row=i, column=1, padx=12, pady=5)
            entries[i] = e
        def do_erase():
            try:
                start = int(entries[0].get()); end = int(entries[1].get())
                if end < start: messagebox.showerror("Error","End ≥ Start required"); return
                if messagebox.askyesno("Confirm Erase",
                    f"Erase sectors {start}–{end}?\nThis CANNOT be undone."):
                    self.ble.send_sd_erase(start, end); dlg.destroy()
            except ValueError: messagebox.showerror("Error","Invalid numbers")
        bf = tk.Frame(dlg, bg='#0d1229'); bf.pack(pady=12)
        self._btn(bf, "Erase", self.accent_orange, 'black',
                  do_erase, w=12, h=2).pack(side=tk.LEFT, padx=5)
        self._btn(bf, "Cancel", self.accent_red, 'white',
                  dlg.destroy, w=10, h=2).pack(side=tk.LEFT, padx=5)

    def _clear_raw(self):
        self.raw_hex_text.delete(1.0, tk.END)
        self._raw_frame_count = 0
        self.raw_frame_count_lbl.config(text="Frames: 0")

    def _toggle_raw_pause(self):
        self._raw_paused = not self._raw_paused
        self.pause_raw_btn.config(text="RESUME" if self._raw_paused else "PAUSE")

    def _add_log(self, message: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_messages.insert(tk.END, f"[{ts}] {message}\n")
        self.log_messages.see(tk.END)

    def _on_close(self):
        print("\n👋 Shutting down…")
        if self.bat_logger.logging_active: self.bat_logger.stop_logging()
        if self.sd_logger.logging_active:  self.sd_logger.stop_logging()
        self.ble.stop()
        self.root.destroy()

    def run(self):
        self.root.mainloop()


# =============================================================================
# ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    print("╔══════════════════════════════════════════════════════════════════════╗")
    print("║  VPUSH BMS DASHBOARD v5.2 — INTEGRATED OTA (FULLY FIXED)            ║")
    print("║  OTA: TRIGGER→FLASH→SIZE→CHUNKS→CRC  (fully automatic)              ║")
    print("║  FIX 1: OTA engine EXACT match to ota.py v3.0                        ║")
    print("║  FIX 2: Proper asyncio event handling for ACKs                       ║")
    print("║  FIX 3: Notification handler sets OTA events on correct loop         ║")
    print("║  CRC-32: reflected poly 0xEDB88320 (matches bsp.c BSP_Crc32_u32)    ║")
    print("╚══════════════════════════════════════════════════════════════════════╝")
    app = VPUSH_BMS_GUI()
    app.run()